"""
MLB Daily Betting Report Generator
====================================
Workflow:
  1. Auto-fetches today's MLB schedule from MLB Stats API (free, no key needed)
  2. Auto-updates team game logs with recent results
  3. Client pastes today's moneylines into: daily_odds_input.xlsx
  4. Script checks each game against all 36 scenarios
  5. Outputs: MLB_Daily_Report_YYYY-MM-DD.xlsx  (Clear Fade + Inconsistent tabs)

Run manually each morning, or schedule via Windows Task Scheduler / cron.
"""

import io
import requests
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
from datetime import datetime, timedelta, date
import re
import os
import sys
import json

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────

DATA_FILE      = 'MLB Data 2023-2026.xlsx'   # historical data (kept updated)
ODDS_INPUT     = 'daily_odds_input.xlsx'      # client fills this each morning
ODDS_CACHE     = 'odds_cache.json'            # persisted moneylines by date/team
OUTPUT_DIR     = '.'                          # where to save daily reports
REPORT_DATE    = date.today()                 # override: date(2026, 7, 7)

# ─────────────────────────────────────────────
# TEAM NAME MAPPING  (MLB API name -> our canonical name)
# ─────────────────────────────────────────────

API_TO_CANONICAL = {
    'Chicago White Sox':    'CHICAGO WHITE SOX',
    'Cleveland Guardians':  'CLEVELAND GUARDIANS',
    'Detroit Tigers':       'DETROIT TIGERS',
    'Kansas City Royals':   'KANSAS CITY ROYALS',
    'Minnesota Twins':      'MINNESOTA TWINS',
    'Baltimore Orioles':    'BALTIMORE ORIOLES',
    'Boston Red Sox':       'BOSTON RED SOX',
    'New York Yankees':     'NEW YORK YANKEES',
    'Tampa Bay Rays':       'TAMPA BAY RAYS',
    'Toronto Blue Jays':    'TORONTO BLUE JAYS',
    'Oakland Athletics':    'ATHLETICS',
    'Athletics':            'ATHLETICS',
    'Houston Astros':       'HOUSTON ASTROS',
    'Los Angeles Angels':   'LOS ANGELES ANGELS',
    'Seattle Mariners':     'SEATTLE MARINERS',
    'Texas Rangers':        'TEXAS RANGERS',
    'Chicago Cubs':         'CHICAGO CUBS',
    'Cincinnati Reds':      'CINCINNATI REDS',
    'Milwaukee Brewers':    'MILWAUKEE BREWERS',
    'Pittsburgh Pirates':   'PITTSBURGH PIRATES',
    'St. Louis Cardinals':  'ST. LOUIS CARDINALS',
    'Atlanta Braves':       'ATLANTA BRAVES',
    'Miami Marlins':        'MIAMI MARLINS',
    'New York Mets':        'NEW YORK METS',
    'Philadelphia Phillies':'PHILADELPHIA PHILLIES',
    'Washington Nationals': 'WASHINGTON NATIONALS',
    'Arizona Diamondbacks': 'ARIZONA DIAMONDBACKS',
    'Colorado Rockies':     'COLORADO ROCKIES',
    'Los Angeles Dodgers':  'LOS ANGELES DODGERS',
    'San Diego Padres':     'SAN DIEGO PADRES',
    'San Francisco Giants': 'SAN FRANCISCO GIANTS',
}

TEAMS = list(set(API_TO_CANONICAL.values()))

AL_EAST    = {'BALTIMORE ORIOLES','BOSTON RED SOX','NEW YORK YANKEES','TAMPA BAY RAYS','TORONTO BLUE JAYS'}
AL_CENTRAL = {'CHICAGO WHITE SOX','CLEVELAND GUARDIANS','DETROIT TIGERS','KANSAS CITY ROYALS','MINNESOTA TWINS'}
AL_WEST    = {'ATHLETICS','HOUSTON ASTROS','LOS ANGELES ANGELS','SEATTLE MARINERS','TEXAS RANGERS'}
NL_EAST    = {'ATLANTA BRAVES','MIAMI MARLINS','NEW YORK METS','PHILADELPHIA PHILLIES','WASHINGTON NATIONALS'}
NL_CENTRAL = {'CHICAGO CUBS','CINCINNATI REDS','MILWAUKEE BREWERS','PITTSBURGH PIRATES','ST. LOUIS CARDINALS'}
NL_WEST    = {'ARIZONA DIAMONDBACKS','COLORADO ROCKIES','LOS ANGELES DODGERS','SAN DIEGO PADRES','SAN FRANCISCO GIANTS'}

DIVISIONS = {}
for t in AL_EAST:    DIVISIONS[t] = 'AL_EAST'
for t in AL_CENTRAL: DIVISIONS[t] = 'AL_CENTRAL'
for t in AL_WEST:    DIVISIONS[t] = 'AL_WEST'
for t in NL_EAST:    DIVISIONS[t] = 'NL_EAST'
for t in NL_CENTRAL: DIVISIONS[t] = 'NL_CENTRAL'
for t in NL_WEST:    DIVISIONS[t] = 'NL_WEST'

OPP_NORM = {
    'houston': 'HOUSTON ASTROS', 'san francisco': 'SAN FRANCISCO GIANTS',
    'pittsburgh': 'PITTSBURGH PIRATES', 'minnesota': 'MINNESOTA TWINS',
    'baltimore': 'BALTIMORE ORIOLES', 'philadelphia': 'PHILADELPHIA PHILLIES',
    'boston': 'BOSTON RED SOX', 'ny yankees': 'NEW YORK YANKEES',
    'new york yankees': 'NEW YORK YANKEES', 'tampa bay': 'TAMPA BAY RAYS',
    'toronto': 'TORONTO BLUE JAYS', 'la angels': 'LOS ANGELES ANGELS',
    'los angeles angels': 'LOS ANGELES ANGELS', 'chicago white sox': 'CHICAGO WHITE SOX',
    'chi white sox': 'CHICAGO WHITE SOX', 'seattle': 'SEATTLE MARINERS',
    'oakland': 'ATHLETICS', 'athletics': 'ATHLETICS',
    'cleveland': 'CLEVELAND GUARDIANS', 'detroit': 'DETROIT TIGERS',
    'kansas city': 'KANSAS CITY ROYALS', 'texas': 'TEXAS RANGERS',
    'chicago cubs': 'CHICAGO CUBS', 'chi cubs': 'CHICAGO CUBS',
    'cincinnati': 'CINCINNATI REDS', 'milwaukee': 'MILWAUKEE BREWERS',
    'st. louis': 'ST. LOUIS CARDINALS', 'atlanta': 'ATLANTA BRAVES',
    'miami': 'MIAMI MARLINS', 'ny mets': 'NEW YORK METS',
    'new york mets': 'NEW YORK METS', 'washington': 'WASHINGTON NATIONALS',
    'arizona': 'ARIZONA DIAMONDBACKS', 'colorado': 'COLORADO ROCKIES',
    'la dodgers': 'LOS ANGELES DODGERS', 'los angeles dodgers': 'LOS ANGELES DODGERS',
    'san diego': 'SAN DIEGO PADRES', 'chicago-al': 'CHICAGO WHITE SOX',
    'chicago-nl': 'CHICAGO CUBS', 'new york-al': 'NEW YORK YANKEES',
    'new york-nl': 'NEW YORK METS',
}

OPP_ABBREV = {
    'sd': 'SAN DIEGO PADRES', 'sf': 'SAN FRANCISCO GIANTS',
    'lad': 'LOS ANGELES DODGERS', 'laa': 'LOS ANGELES ANGELS', 'ana': 'LOS ANGELES ANGELS',
    'nyy': 'NEW YORK YANKEES', 'nym': 'NEW YORK METS',
    'chc': 'CHICAGO CUBS', 'cws': 'CHICAGO WHITE SOX', 'chw': 'CHICAGO WHITE SOX',
    'bos': 'BOSTON RED SOX', 'tb': 'TAMPA BAY RAYS', 'tba': 'TAMPA BAY RAYS',
    'tor': 'TORONTO BLUE JAYS', 'bal': 'BALTIMORE ORIOLES', 'cle': 'CLEVELAND GUARDIANS',
    'det': 'DETROIT TIGERS', 'kc': 'KANSAS CITY ROYALS', 'kcr': 'KANSAS CITY ROYALS',
    'min': 'MINNESOTA TWINS', 'hou': 'HOUSTON ASTROS', 'sea': 'SEATTLE MARINERS',
    'tex': 'TEXAS RANGERS', 'oak': 'ATHLETICS', 'ath': 'ATHLETICS',
    'atl': 'ATLANTA BRAVES', 'mia': 'MIAMI MARLINS', 'phi': 'PHILADELPHIA PHILLIES',
    'was': 'WASHINGTON NATIONALS', 'wsh': 'WASHINGTON NATIONALS',
    'cin': 'CINCINNATI REDS', 'mil': 'MILWAUKEE BREWERS', 'pit': 'PITTSBURGH PIRATES',
    'stl': 'ST. LOUIS CARDINALS', 'ari': 'ARIZONA DIAMONDBACKS', 'az': 'ARIZONA DIAMONDBACKS',
    'col': 'COLORADO ROCKIES',
}

def normalize_opponent(raw_opp, warnings=None):
    s = str(raw_opp).replace('\xa0', ' ').strip()
    s = re.sub(r'^(vs|@)\s*', '', s, flags=re.IGNORECASE).strip()
    sl = s.lower()
    if sl == 'los angeles' and warnings is not None:
        warnings.append(
            f'Ambiguous opponent "{s}" — use "Los Angeles Dodgers" or "Los Angeles Angels".'
        )
    if sl in OPP_NORM:
        return OPP_NORM[sl]
    if sl in OPP_ABBREV:
        return OPP_ABBREV[sl]
    for key in sorted(OPP_NORM.keys(), key=len, reverse=True):
        if sl == key or sl.startswith(key + ' ') or sl.endswith(' ' + key):
            return OPP_NORM[key]
    return s.upper()

def parse_score(score_str, result, warnings=None):
    try:
        parts = str(score_str).split('-')
        a, b = int(parts[0]), int(parts[1])
        return (a, b) if result == 'W' else (b, a)
    except (TypeError, ValueError, IndexError):
        if warnings is not None:
            warnings.append(f'Could not parse score "{score_str}" — run-window scenarios may not fire.')
        return None, None

# ─────────────────────────────────────────────
# STEP 1: LOAD HISTORICAL DATA
# ─────────────────────────────────────────────

def load_historical_data():
    wb = load_workbook(DATA_FILE, read_only=True)
    records = []
    TEAM_LIST = [
        'CHICAGO WHITE SOX','CLEVELAND GUARDIANS','DETROIT TIGERS','KANSAS CITY ROYALS','MINNESOTA TWINS',
        'BALTIMORE ORIOLES','BOSTON RED SOX','NEW YORK YANKEES','TAMPA BAY RAYS','TORONTO BLUE JAYS',
        'ATHLETICS','HOUSTON ASTROS','LOS ANGELES ANGELS','SEATTLE MARINERS','TEXAS RANGERS',
        'CHICAGO CUBS','CINCINNATI REDS','MILWAUKEE BREWERS','PITTSBURGH PIRATES','ST. LOUIS CARDINALS',
        'ATLANTA BRAVES','MIAMI MARLINS','NEW YORK METS','PHILADELPHIA PHILLIES','WASHINGTON NATIONALS',
        'ARIZONA DIAMONDBACKS','COLORADO ROCKIES','LOS ANGELES DODGERS','SAN DIEGO PADRES','SAN FRANCISCO GIANTS'
    ]
    for year_str in ['2023','2024','2025','2026']:
        ws = wb[year_str]
        all_rows = list(ws.iter_rows(values_only=True))
        year = int(year_str)
        for t_idx, team in enumerate(TEAM_LIST):
            col = t_idx * 9
            for row in all_rows[2:]:
                date_val = row[col]
                opp_val  = row[col+1]
                result   = row[col+3]
                score    = row[col+4]
                line     = row[col+5]
                ou       = row[col+6]
                total    = row[col+7]
                if date_val is None or result is None:
                    continue
                if str(result).strip().upper() not in ('W', 'L'):
                    continue
                opp_str = str(opp_val).replace('\xa0',' ')
                home_away = 'away' if opp_str.startswith('@') else 'home'
                opponent = normalize_opponent(opp_str)
                rs, ra = parse_score(score, result)
                records.append({
                    'year': year, 'team': team,
                    'date': pd.Timestamp(date_val),
                    'opponent': opponent, 'home_away': home_away,
                    'result': result, 'score': score,
                    'runs_scored': rs, 'runs_allowed': ra,
                    'line': line, 'ou': ou, 'total': total,
                    'division': DIVISIONS.get(team,'UNKNOWN'),
                })
    df = pd.DataFrame(records)
    df['date'] = pd.to_datetime(df['date'])
    return df.sort_values(['team','date']).reset_index(drop=True)

# ─────────────────────────────────────────────
# STEP 2: FETCH TODAY'S SCHEDULE FROM MLB API
# ─────────────────────────────────────────────

def fetch_todays_schedule(report_date):
    date_str = report_date.strftime('%Y-%m-%d')
    url = f'https://statsapi.mlb.com/api/v1/schedule?sportId=1&date={date_str}'
    try:
        r = requests.get(url, timeout=10)
        data = r.json()
    except Exception as e:
        print(f'ERROR fetching schedule: {e}')
        return []

    games = []
    if not data.get('dates'):
        print(f'No games found for {date_str}')
        return []

    for g in data['dates'][0]['games']:
        away_api = g['teams']['away']['team']['name']
        home_api = g['teams']['home']['team']['name']
        away = API_TO_CANONICAL.get(away_api, away_api.upper())
        home = API_TO_CANONICAL.get(home_api, home_api.upper())
        game_time = g.get('gameDate','')
        games.append({
            'away_team': away,
            'home_team': home,
            'game_time': game_time,
            'game_pk':   g['gamePk'],
        })
    print(f'Schedule: {len(games)} games on {date_str}')
    return games

# ─────────────────────────────────────────────
# ODDS CACHE (persist moneylines for API-fetched games)
# ─────────────────────────────────────────────

def load_odds_cache():
    if not os.path.exists(ODDS_CACHE):
        return {}
    try:
        with open(ODDS_CACHE, encoding='utf-8') as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}


def _coerce_line(line):
    if line is None:
        return None
    try:
        return int(float(line))
    except (TypeError, ValueError):
        return None


def save_odds_to_cache(report_date, odds_dict):
    """Store entered moneylines so API-fetched results can backfill lines later."""
    if not odds_dict:
        return
    cache = load_odds_cache()
    key = report_date.strftime('%Y-%m-%d') if hasattr(report_date, 'strftime') else str(report_date)[:10]
    day = cache.setdefault(key, {})
    for team, line in odds_dict.items():
        ln = _coerce_line(line)
        if ln is not None:
            day[str(team).upper()] = ln
    with open(ODDS_CACHE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, indent=2)


def _line_from_odds_cache(cache, dt, team):
    d = dt.strftime('%Y-%m-%d') if hasattr(dt, 'strftime') else str(dt)[:10]
    raw = cache.get(d, {}).get(str(team).upper())
    return _coerce_line(raw)


def apply_odds_cache(df):
    """Backfill missing lines on historical/API rows from the odds cache."""
    cache = load_odds_cache()
    if not cache:
        return df
    for idx, row in df.iterrows():
        if _coerce_line(row.get('line')) is not None:
            continue
        ln = _line_from_odds_cache(cache, row['date'], row['team'])
        if ln is not None:
            df.at[idx, 'line'] = ln
    return df


def get_team_winpcts_as_of(report_date):
    """
    Season win% for each MLB API team name as of report_date
    (games completed strictly before that date).
    """
    df = load_historical_data()
    df, _ = fetch_recent_results(df, report_date)
    before = df[df['date'].dt.date < report_date]
    by_canonical = {}
    for team in set(API_TO_CANONICAL.values()):
        tdf = before[before['team'] == team]
        w = (tdf['result'] == 'W').sum()
        l = (tdf['result'] == 'L').sum()
        total = w + l
        by_canonical[team] = round(w / total, 4) if total > 0 else 0.500
    return {api: by_canonical.get(canon, 0.500) for api, canon in API_TO_CANONICAL.items()}

# ─────────────────────────────────────────────
# STEP 3: FETCH RECENT RESULTS & UPDATE DATA
# ─────────────────────────────────────────────

def fetch_recent_results(df, report_date):
    """
    Fetch game results from MLB Stats API for dates after the last date in df.
    Appends new completed games to df.

    Returns:
        (updated_df, warnings) — warnings lists dates that could not be fetched.
    """
    last_date = df['date'].max().date()
    check_date = last_date + timedelta(days=1)
    new_records = []
    warnings = []

    while check_date < report_date:
        date_str = check_date.strftime('%Y-%m-%d')
        url = f'https://statsapi.mlb.com/api/v1/schedule?sportId=1&date={date_str}&hydrate=linescore'
        try:
            r = requests.get(url, timeout=10)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            warnings.append(f'Could not fetch results for {date_str}: {e}')
            check_date += timedelta(days=1)
            continue

        if not data.get('dates'):
            check_date += timedelta(days=1)
            continue

        day_games = [g for g in data['dates'][0]['games'] if g['status']['detailedState'] == 'Final']
        for g in day_games:
            away_api   = g['teams']['away']['team']['name']
            home_api   = g['teams']['home']['team']['name']
            away_team  = API_TO_CANONICAL.get(away_api, away_api.upper())
            home_team  = API_TO_CANONICAL.get(home_api, home_api.upper())
            away_score = g['teams']['away'].get('score', 0)
            home_score = g['teams']['home'].get('score', 0)
            if away_score == home_score:
                continue
            away_win = g['teams']['away'].get('isWinner', False)
            home_win = g['teams']['home'].get('isWinner', False)
            if away_win:
                rs_a, ra_a, res_a = away_score, home_score, 'W'
                rs_h, ra_h, res_h = home_score, away_score, 'L'
            elif home_win:
                rs_a, ra_a, res_a = away_score, home_score, 'L'
                rs_h, ra_h, res_h = home_score, away_score, 'W'
            else:
                continue

            score_str = f'{max(away_score, home_score)}-{min(away_score, home_score)}'
            odds_cache = load_odds_cache()

            year = check_date.year
            dt   = pd.Timestamp(check_date)

            for team, opp, ha, rs, ra, res in [
                (away_team, home_team, 'away', rs_a, ra_a, res_a),
                (home_team, away_team, 'home', rs_h, ra_h, res_h),
            ]:
                cached_line = _line_from_odds_cache(odds_cache, dt, team)
                new_records.append({
                    'year': year, 'team': team,
                    'date': dt, 'opponent': opp,
                    'home_away': ha, 'result': res,
                    'score': score_str,
                    'runs_scored': rs, 'runs_allowed': ra,
                    'line': cached_line, 'ou': None, 'total': None,
                    'division': DIVISIONS.get(team, 'UNKNOWN'),
                })

        check_date += timedelta(days=1)

    if new_records:
        df = _append_game_records(df, new_records)
        df = df.sort_values(['team','date']).reset_index(drop=True)
        print(f'Added {len(new_records)} new game records from API')
    else:
        print('Data already up to date')

    df = apply_odds_cache(df)
    return df, warnings

# ─────────────────────────────────────────────
# STEP 4: COMPUTE TEAM STATES
# ─────────────────────────────────────────────

def compute_all_states(df):
    TEAM_LIST = sorted(df['team'].unique())
    parts = []
    all_cols = list(df.columns)
    for team in TEAM_LIST:
        tdf = df[df['team'] == team].copy().sort_values('date').reset_index(drop=True)
        parts.append(compute_team_states(tdf))
    if not parts:
        return df.iloc[0:0].copy()
    enriched = parts[0]
    for part in parts[1:]:
        enriched = pd.concat([enriched, part.reindex(columns=enriched.columns)], ignore_index=True)
    return enriched

def compute_team_states(rows):
    n = len(rows)
    streak = [0]*n
    prev_opponent = [None]*n
    prev_result = [None]*n
    prev_line = [None]*n
    prev_runs_scored = [None]*n
    prev_runs_allowed = [None]*n
    series_game_num = [1]*n
    series_id = [0]*n
    homestand_game_num = [0]*n
    roadtrip_game_num = [0]*n
    homestand_series_num = [0]*n
    wins_so_far = [0]*n
    games_so_far = [0]*n
    prev2_rs = [None]*n
    prev3_rs = [None]*n
    prev4_rs = [None]*n
    prev2_ra = [None]*n
    prev3_ra = [None]*n
    prev4_ra = [None]*n

    cur_streak = 0; cur_wins = 0; cur_games = 0; s_id = 0
    runs_scored_hist = []; runs_allowed_hist = []
    cur_home_game = 0; cur_road_game = 0; cur_home_series = 0
    last_location = None; cur_series_start = 0
    cur_year = rows.at[0,'year'] if n > 0 else None

    for i in range(n):
        this_year = rows.at[i,'year']
        if this_year != cur_year:
            cur_streak=0; cur_wins=0; cur_games=0
            runs_scored_hist=[]; runs_allowed_hist=[]
            cur_home_game=0; cur_road_game=0; cur_home_series=0
            last_location=None; cur_year=this_year
            prev_result[i]=None; prev_opponent[i]=None
            prev_line[i]=None; prev_runs_scored[i]=None; prev_runs_allowed[i]=None

        wins_so_far[i] = cur_wins
        games_so_far[i] = cur_games

        if i > 0 and prev_result[i] is None:
            prev_result[i]      = rows.at[i-1,'result']
            prev_opponent[i]    = rows.at[i-1,'opponent']
            prev_line[i]        = rows.at[i-1,'line']
            prev_runs_scored[i] = rows.at[i-1,'runs_scored']
            prev_runs_allowed[i]= rows.at[i-1,'runs_allowed']

        streak[i] = cur_streak

        opp = rows.at[i,'opponent']
        if i == 0 or opp != rows.at[i-1,'opponent']:
            s_id += 1; cur_series_start = i
        series_id[i] = s_id
        series_game_num[i] = i - cur_series_start + 1

        loc = rows.at[i,'home_away']
        if loc != last_location:
            cur_home_game=0; cur_road_game=0; cur_home_series=0; last_location=loc
        if loc == 'home':
            cur_home_game += 1
            if i == 0 or rows.at[i-1,'home_away'] != 'home' or opp != rows.at[i-1,'opponent']:
                cur_home_series = cur_home_series+1 if (i>0 and rows.at[i-1,'home_away']=='home') else 1
            homestand_game_num[i] = cur_home_game
            homestand_series_num[i] = cur_home_series
            roadtrip_game_num[i] = 0
        else:
            cur_road_game += 1
            roadtrip_game_num[i] = cur_road_game
            homestand_game_num[i] = 0
            homestand_series_num[i] = 0

        if len(runs_scored_hist) >= 2:
            prev2_rs[i] = list(runs_scored_hist[-2:])
            prev2_ra[i] = list(runs_allowed_hist[-2:])
        if len(runs_scored_hist) >= 3:
            prev3_rs[i] = list(runs_scored_hist[-3:])
            prev3_ra[i] = list(runs_allowed_hist[-3:])
        if len(runs_scored_hist) >= 4:
            prev4_rs[i] = list(runs_scored_hist[-4:])
            prev4_ra[i] = list(runs_allowed_hist[-4:])

        r = rows.at[i,'result']
        runs_scored_hist.append(rows.at[i,'runs_scored'])
        runs_allowed_hist.append(rows.at[i,'runs_allowed'])
        cur_games += 1
        if r == 'W':
            cur_wins += 1
            cur_streak = cur_streak+1 if cur_streak >= 0 else 1
        elif r == 'L':
            cur_streak = cur_streak-1 if cur_streak <= 0 else -1

    sbp = [None] + streak[:-1]

    rows['streak_before']       = streak
    rows['streak_before_prev']  = sbp
    rows['prev_opponent']       = prev_opponent
    rows['prev_result']         = prev_result
    rows['prev_line']           = prev_line
    rows['prev_runs_scored']    = prev_runs_scored
    rows['prev_runs_allowed']   = prev_runs_allowed
    rows['series_game_num']     = series_game_num
    rows['series_id']           = series_id
    rows['homestand_game_num']  = homestand_game_num
    rows['homestand_series_num']= homestand_series_num
    rows['roadtrip_game_num']   = roadtrip_game_num
    rows['wins_before']         = wins_so_far
    rows['games_before']        = games_so_far
    rows['prev2_runs_scored']   = prev2_rs
    rows['prev3_runs_scored']   = prev3_rs
    rows['prev4_runs_scored']   = prev4_rs
    rows['prev2_runs_allowed']  = prev2_ra
    rows['prev3_runs_allowed']  = prev3_ra
    rows['prev4_runs_allowed']  = prev4_ra

    rows['winpct_before'] = rows.apply(
        lambda r: r['wins_before']/r['games_before'] if r['games_before']>0 else None, axis=1)

    results_list = rows['result'].tolist()
    year_list    = rows['year'].tolist()
    l10 = []
    for i in range(len(results_list)):
        window = []
        for j in range(i-1, max(-1,i-11), -1):
            if year_list[j] != year_list[i]: break
            window.append(results_list[j])
        l10.append(window.count('W'))
    rows['last10_wins'] = l10

    rows['series_total'] = _series_length_totals(rows)
    rows['rt_segment'] = (rows['home_away'] != rows['home_away'].shift()).cumsum()
    away_trip_totals = _segment_length_totals(rows, 'away')
    home_trip_totals = _segment_length_totals(rows, 'home')
    rows['roadtrip_total'] = 0
    rows['homestand_total'] = 0
    for i in range(len(rows)):
        if rows.iloc[i]['home_away'] == 'away':
            rows.iat[i, rows.columns.get_loc('roadtrip_total')] = away_trip_totals[i]
        else:
            rows.iat[i, rows.columns.get_loc('homestand_total')] = home_trip_totals[i]

    return rows


def _series_length_totals(rows):
    """
    Full series length once a series ends (next game is vs a different opponent).
    Open series at end of log use seg_len+2 so last_game_series does not fire early.
    """
    n = len(rows)
    totals = [0] * n
    i = 0
    while i < n:
        opp = rows.iloc[i]['opponent']
        start = i
        while i < n and rows.iloc[i]['opponent'] == opp:
            i += 1
        end = i - 1
        seg_len = end - start + 1
        total_val = seg_len if i < n else seg_len + 2
        for j in range(start, end + 1):
            totals[j] = total_val
    return totals


def _segment_length_totals(rows, location):
    """Same as _series_length_totals but for consecutive home or away segments."""
    n = len(rows)
    totals = [0] * n
    i = 0
    while i < n:
        if rows.iloc[i]['home_away'] != location:
            i += 1
            continue
        start = i
        while i < n and rows.iloc[i]['home_away'] == location:
            i += 1
        end = i - 1
        seg_len = end - start + 1
        total_val = seg_len if i < n else seg_len + 2
        for j in range(start, end + 1):
            totals[j] = total_val
    return totals


def _append_game_records(df, new_records):
    """Append API game rows without pandas concat dtype warnings."""
    if not new_records:
        return df
    new_df = pd.DataFrame(new_records)
    combined_cols = list(dict.fromkeys(list(df.columns) + list(new_df.columns)))
    df = df.reindex(columns=combined_cols)
    new_df = new_df.reindex(columns=combined_cols)
    return pd.concat([df, new_df], ignore_index=True)


def _run_windows_from_history(tdf):
    """Last 2/3/4 game run windows including the most recent completed game."""
    rs = tdf['runs_scored'].tolist()
    ra = tdf['runs_allowed'].tolist()

    def take(hist, k):
        return hist[-k:] if len(hist) >= k else None

    return {
        'prev2_runs_scored': take(rs, 2),
        'prev3_runs_scored': take(rs, 3),
        'prev4_runs_scored': take(rs, 4),
        'prev2_runs_allowed': take(ra, 2),
        'prev3_runs_allowed': take(ra, 3),
        'prev4_runs_allowed': take(ra, 4),
    }


def _last10_wins_from_history(tdf):
    """Wins in the last 10 completed games (same season year as most recent game)."""
    last_year = tdf.iloc[-1]['year']
    recent = tdf[tdf['year'] == last_year].tail(10)
    return int((recent['result'] == 'W').sum())


def build_streak_lookup(df):
    """Point-in-time streak_before keyed by (date, team)."""
    return df.set_index(['date', 'team'])['streak_before'].to_dict()


def build_opp_road_wpct_lookup(df):
    """Opponent road win% before each game, keyed by (date, team)."""
    lookup = {}
    for team, tdf in df.groupby('team'):
        tdf = tdf.sort_values('date')
        road = tdf[tdf['home_away'] == 'away']
        for _, row in tdf.iterrows():
            d = row['date']
            rb = road[road['date'] < d]
            if rb.empty:
                lookup[(d, team)] = None
            else:
                rw = (rb['result'] == 'W').sum()
                rl = (rb['result'] == 'L').sum()
                lookup[(d, team)] = rw / (rw + rl) if (rw + rl) > 0 else None
    return lookup


def backtest_wl_counts(subset, verdict):
    """Historical backtest W/L: fade wins when the faded team loses."""
    if verdict == 'INCONSISTENT':
        return None, None
    if verdict == 'CLEAR FADE':
        wins = int((subset['result'] == 'L').sum())
        losses = int((subset['result'] == 'W').sum())
    else:
        wins = int((subset['result'] == 'W').sum())
        losses = int((subset['result'] == 'L').sum())
    return wins, losses


def load_master_history_before(report_date):
    """Rows from Master_Results.xlsx with Date strictly before report_date."""
    from master_results_manager import MASTER_FILE, parse_results_upload
    if not os.path.exists(MASTER_FILE):
        return []
    with open(MASTER_FILE, 'rb') as f:
        df, _, _, _ = parse_results_upload(f.read())
    day = report_date.strftime('%Y-%m-%d') if hasattr(report_date, 'strftime') else str(report_date)[:10]
    return [row for _, row in df.iterrows() if str(row.get('Date', ''))[:10] < day]


# ─────────────────────────────────────────────
# STEP 5: GET CURRENT STATE FOR TODAY'S GAMES
# ─────────────────────────────────────────────

def get_team_state(enriched, team, report_date):
    """Get the most recent state for a team as of report_date (before today's game)."""
    tdf = enriched[
        (enriched['team'] == team) &
        (enriched['date'].dt.date < report_date)
    ].sort_values('date')

    if tdf.empty:
        return None

    last = tdf.iloc[-1]
    # Compute current streak after last game
    sb = last['streak_before']
    res = last['result']
    if res == 'W':
        cur_streak = sb + 1 if sb >= 0 else 1
    else:
        cur_streak = sb - 1 if sb <= 0 else -1

    run_windows = _run_windows_from_history(tdf)

    return {
        'team':               team,
        'streak_before':      cur_streak,
        'streak_before_prev': last['streak_before'],
        'prev_result':        last['result'],
        'prev_opponent':      last['opponent'],
        'prev_line':          last['line'],
        'prev_runs_scored':   last['runs_scored'],
        'prev_runs_allowed':  last['runs_allowed'],
        'series_game_num':    last['series_game_num'],
        'series_total':       last['series_total'],
        'series_id':          last['series_id'],
        'homestand_game_num': last['homestand_game_num'],
        'homestand_series_num': last['homestand_series_num'],
        'roadtrip_game_num':  last['roadtrip_game_num'],
        'roadtrip_total':     last['roadtrip_total'],
        'homestand_total':    last['homestand_total'],
        'wins_before':        last['wins_before'] + (1 if last['result']=='W' else 0),
        'games_before':       last['games_before'] + 1,
        'last10_wins':        _last10_wins_from_history(tdf),
        'prev2_runs_scored':  run_windows['prev2_runs_scored'],
        'prev3_runs_scored':  run_windows['prev3_runs_scored'],
        'prev4_runs_scored':  run_windows['prev4_runs_scored'],
        'prev2_runs_allowed': run_windows['prev2_runs_allowed'],
        'prev3_runs_allowed': run_windows['prev3_runs_allowed'],
        'prev4_runs_allowed': run_windows['prev4_runs_allowed'],
        'winpct_before':      (last['wins_before']+(1 if last['result']=='W' else 0)) / (last['games_before']+1),
        'division':           DIVISIONS.get(team,'UNKNOWN'),
    }

def build_game_row(state, home_away, opponent, line):
    """Build a pseudo-row for scenario filtering from team state + today's game info."""
    if state is None:
        return None

    # Determine series position for today's game
    last_opp = state['prev_opponent']
    if last_opp == opponent:
        series_game = state['series_game_num'] + 1
        series_total = max(state['series_total'], series_game)
    else:
        series_game = 1
        series_total = 3

    # Homestand/road trip position
    if home_away == 'home':
        hg = state['homestand_game_num'] + 1 if state['homestand_game_num'] > 0 else 1
        rg = 0
        hs_series = state['homestand_series_num']
        if last_opp != opponent and state['homestand_game_num'] > 0:
            hs_series += 1
        elif state['homestand_game_num'] == 0:
            hs_series = 1
        rt_total = 0
    else:
        if state['roadtrip_game_num'] > 0 and state['roadtrip_game_num'] < state['roadtrip_total']:
            rg = state['roadtrip_game_num'] + 1
            rt_total = max(state['roadtrip_total'], rg)
        else:
            rg = 1
            rt_total = 1
        hg = 0
        hs_series = 0

    return {
        'team':               state['team'],
        'opponent':           opponent,
        'home_away':          home_away,
        'line':               line,
        'streak_before':      state['streak_before'],
        'streak_before_prev': state['streak_before_prev'],
        'prev_result':        state['prev_result'],
        'prev_opponent':      state['prev_opponent'],
        'prev_line':          state['prev_line'],
        'prev_runs_scored':   state['prev_runs_scored'],
        'prev_runs_allowed':  state['prev_runs_allowed'],
        'series_game_num':    series_game,
        'series_total':       series_total,
        'homestand_game_num': hg,
        'homestand_series_num': hs_series,
        'roadtrip_game_num':  rg,
        'roadtrip_total':     rt_total,
        'last10_wins':        state['last10_wins'],
        'wins_before':        state['wins_before'],
        'games_before':       state['games_before'],
        'winpct_before':      state['winpct_before'],
        'prev2_runs_scored':  state['prev2_runs_scored'],
        'prev3_runs_scored':  state['prev3_runs_scored'],
        'prev4_runs_scored':  state['prev4_runs_scored'],
        'prev2_runs_allowed': state['prev2_runs_allowed'],
        'prev3_runs_allowed': state['prev3_runs_allowed'],
        'prev4_runs_allowed': state['prev4_runs_allowed'],
        'division':           state['division'],
    }


# ─────────────────────────────────────────────
# STEP 6: SCENARIO FILTERS (all 36)
# ─────────────────────────────────────────────

def same_opp(r): return r['prev_opponent'] == r['opponent']
def diff_opp(r): return r['prev_opponent'] != r['opponent'] and r['prev_opponent'] is not None
def is_home(r):  return r['home_away'] == 'home'
def is_away(r):  return r['home_away'] == 'away'
def prev_won(r): return r['prev_result'] == 'W'
def prev_lost(r):return r['prev_result'] == 'L'

def run_diff_prev(r):
    rs, ra = r['prev_runs_scored'], r['prev_runs_allowed']
    return (rs-ra) if (rs is not None and ra is not None) else None

def all_rs_le(lst, t):
    return lst is not None and all(v is not None and v<=t for v in lst)

def all_ra_ge(lst, t):
    return lst is not None and all(v is not None and v>=t for v in lst)

def last_game_series(r): return r['series_game_num'] == r['series_total']
def first_game_series(r): return r['series_game_num'] == 1
def second_game_series(r): return r['series_game_num'] == 2
def last_game_roadtrip(r): return is_away(r) and r['roadtrip_game_num']==r['roadtrip_total'] and r['roadtrip_total']>0
def first_game_homestand(r): return is_home(r) and r['homestand_game_num']==1
def div_match(t1,t2): return DIVISIONS.get(t1)==DIVISIONS.get(t2)

def s01(r):
    if not prev_won(r) or not same_opp(r): return False
    d = run_diff_prev(r)
    return d is not None and d >= 8

def s02(r):
    if not prev_won(r) or not diff_opp(r): return False
    d = run_diff_prev(r)
    return d is not None and d >= 7

def s03(r):
    if not prev_won(r) or not diff_opp(r): return False
    d = run_diff_prev(r)
    return d is not None and d >= 11

def s04(r):
    pl, cl = r['prev_line'], r['line']
    if pl is None or cl is None: return False
    return prev_won(r) and same_opp(r) and pl >= 150 and cl <= 125

def s05(r):
    pl, cl = r['prev_line'], r['line']
    if pl is None or cl is None: return False
    return prev_lost(r) and diff_opp(r) and pl <= -150 and cl >= 100

def s06(r):
    if not is_away(r) or not prev_lost(r): return False
    sbp = r['streak_before_prev']
    return sbp is not None and sbp >= 5

def s07(r):
    cl = r['line']
    if cl is None: return False
    sbp = r['streak_before_prev']
    return prev_won(r) and diff_opp(r) and sbp is not None and sbp <= -6 and cl <= 175

def s08(r):
    cl = r['line']
    if cl is None: return False
    sbp = r['streak_before_prev']
    return is_home(r) and prev_won(r) and same_opp(r) and sbp is not None and sbp <= -4 and cl >= -200

def s09(r, opp_streaks):
    if not is_away(r) or r['streak_before'] >= 0: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and os >= 4

def s10(r, opp_streaks):
    cl = r['line']
    if cl is None or not is_home(r) or r['streak_before'] >= 0: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and -3 <= os <= -1 and cl <= -150

def s11(r, opp_streaks):
    if not is_home(r) or r['streak_before'] > -4: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and -3 <= os <= -1

def s12(r, opp_streaks):
    if not is_away(r) or r['streak_before'] <= 0: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and os >= 4

def s13(r, opp_streaks):
    if not is_home(r) or r['streak_before'] not in [1,2]: return False
    os = opp_streaks.get(r['opponent'])
    return os is not None and os >= 4

def s14(r):
    cl = r['line']
    if cl is None or not is_away(r) or r['streak_before'] > -2: return False
    return all_rs_le(r['prev2_runs_scored'], 2) and cl >= 150

def s15(r):
    cl = r['line']
    if cl is None or not is_away(r) or r['streak_before'] > -4: return False
    return all_rs_le(r['prev4_runs_scored'], 3) and cl >= -105

def s16(r):
    cl = r['line']
    if cl is None or not is_home(r) or r['streak_before'] > -4: return False
    return all_rs_le(r['prev4_runs_scored'], 2) and cl >= -200

def s17(r):
    if not is_away(r) or r['streak_before'] > -2: return False
    return all_ra_ge(r['prev2_runs_allowed'], 9)

def s18(r):
    cl = r['line']
    if cl is None or not is_away(r) or r['streak_before'] > -3: return False
    return all_ra_ge(r['prev3_runs_allowed'], 7) and cl <= 200

def s19(r):
    if not is_away(r) or r['streak_before'] > -4: return False
    return all_ra_ge(r['prev4_runs_allowed'], 6)

def s20(r): return r['streak_before'] >= 3 and diff_opp(r)
def s21(r): return r['streak_before'] <= -3 and last_game_series(r)

def s22(r):
    cl = r['line']
    return cl is not None and is_away(r) and first_game_series(r) and -130 <= cl <= -111

def s23(r):
    cl = r['line']
    return cl is not None and is_away(r) and last_game_series(r) and 101 <= cl <= 187

def s24(r):
    cl = r['line']
    return cl is not None and last_game_roadtrip(r) and 101 <= cl <= 187

def s25(r):
    cl = r['line']
    return (cl is not None and is_home(r) and first_game_series(r)
            and r['homestand_series_num'] >= 2 and -109 <= cl <= 120)

def s26(r):
    cl = r['line']
    return cl is not None and first_game_homestand(r) and -180 <= cl <= -111

def s27(r):
    cl = r['line']
    return (cl is not None and is_home(r) and last_game_series(r)
            and div_match(r['team'], r['opponent']) and cl <= -180)

def s28(r):
    cl = r['line']
    prs = r['prev_runs_scored']
    return (cl is not None and is_away(r) and first_game_series(r)
            and prev_won(r) and prs is not None and prs >= 6 and cl <= 174)

def s29(r):
    prs = r['prev_runs_scored']
    return (is_away(r) and last_game_series(r) and prev_lost(r)
            and prs is not None and prs <= 4)

def s30(r):
    cl = r['line']
    wp = r['winpct_before']
    return (cl is not None and first_game_homestand(r)
            and wp is not None and wp < 0.500 and 110 <= cl <= 170)

def s31(r):
    cl = r['line']
    prs, pra = r['prev_runs_scored'], r['prev_runs_allowed']
    if cl is None or not prev_lost(r) or not same_opp(r): return False
    if prs is None or pra is None: return False
    return (pra - prs) == 1 and 100 <= cl <= 150

def s32(r, opp_road_wpct):
    cl = r['line']
    if cl is None or not is_home(r) or not prev_lost(r): return False
    if r['last10_wins'] > 5 or cl < 100: return False
    if not div_match(r['team'], r['opponent']): return False
    owp = opp_road_wpct.get(r['opponent'])
    return owp is not None and owp <= 0.400

def s33(r):
    cl, pl = r['line'], r['prev_line']
    if cl is None or pl is None: return False
    if not is_home(r) or not second_game_series(r): return False
    wp = r['winpct_before']
    return (cl >= 100 and wp is not None and wp < 0.500
            and r['prev_result'] == 'L' and pl < 0)

def s34(r):
    cl = r['line']
    if cl is None or not is_home(r): return False
    gb, wb = r['games_before'], r['wins_before']
    if gb == 0 or wb*2 != gb: return False
    return r['streak_before'] >= 2 and cl >= -190

def s35(r):
    cl, pl = r['line'], r['prev_line']
    if cl is None or pl is None: return False
    return second_game_series(r) and cl < 0 and r['prev_result']=='W' and pl < 0

def s36(r):
    if r['series_total'] != 3 or not last_game_series(r): return False
    return r['prev_result'] == 'L' and r['streak_before'] == -1


SCENARIO_DEFS = [
    ('01','BLOWOUT #1 - MJ',                           'CLEAR BET',   s01),
    ('02','BLOWOUT #2 - MJ',                           'CLEAR FADE',  s02),
    ('03','BLOWOUT #3 - MJ',                           'INCONSISTENT',s03),
    ('04','THE BIG UPSET #1 - MJ',                     'INCONSISTENT',s04),
    ('05','THE BIG UPSET #2 - MJ',                     'CLEAR FADE',  s05),
    ('06','THE SNAPPED WINNING STREAK - MJ',           'CLEAR FADE',  s06),
    ('07','THE SNAPPED LOSING STREAK #1 - MJ',         'INCONSISTENT',s07),
    ('08','THE SNAPPED LOSING STREAK #2 - MJ',         'INCONSISTENT',s08),
    ('09','THE COLD TEAM VS HOT TEAM - MJ',            'INCONSISTENT',s09),
    ('10','THE COLD TEAMS MATCHUP #1 - MJ',            'INCONSISTENT',s10),
    ('11','THE COLD TEAMS MATCHUP #2 - MJ',            'INCONSISTENT',s11),
    ('12','THE HOT TEAMS MATCHUP #1 - MJ',             'CLEAR FADE',  s12),
    ('13','THE HOT TEAMS MATCHUP #2 - MJ',             'INCONSISTENT',s13),
    ('14','THE SCORING DROUGHT #1 - MJ',               'CLEAR FADE',  s14),
    ('15','THE SCORING DROUGHT #2 - MJ',               'CLEAR FADE',  s15),
    ('16','THE SCORING DROUGHT #3 - MJ',               'CLEAR FADE',  s16),
    ('17','THE PUMMELED PITCHERS #1 - MJ',             'INCONSISTENT',s17),
    ('18','THE PUMMELED PITCHERS #2 - MJ',             'CLEAR FADE',  s18),
    ('19','THE PUMMELED PITCHERS #3 - MJ',             'CLEAR FADE',  s19),
    ('20','3-GAME WIN STREAK / NEW SERIES - SM',       'CLEAR BET',   s20),
    ('21','3-GAME LOSING STREAK / LAST SERIES - SM',   'CLEAR BET',   s21),
    ('22','SMALL ROAD FAVORITE / NEW SERIES - SM',     'INCONSISTENT',s22),
    ('23','ROAD DOG / LAST GAME OF SERIES - SM',       'INCONSISTENT',s23),
    ('24','ROAD DOG / LAST GAME OF ROAD TRIP - SM',    'CLEAR BET',   s24),
    ('25','HOME DOG / MID-HOMESTAND NEW OPP - SM',     'CLEAR FADE',  s25),
    ('26','SMALL HOME FAV / FIRST HOMESTAND - SM',     'CLEAR FADE',  s26),
    ('27','BIG HOME FAV / LAST DIV SERIES - SM',       'INCONSISTENT',s27),
    ('28','HOT ROAD TEAM / FIRST GAME SERIES - SM',    'INCONSISTENT',s28),
    ('29','COLD ROAD TEAM / LAST GAME SERIES - SM',    'INCONSISTENT',s29),
    ('30','HOME DOG OFF LONG ROAD TRIP',               'CLEAR FADE',  s30),
    ('31','DOG OFF CLOSE LOSS',                        'INCONSISTENT',s31),
    ('32','COLD HOME DOG / DIV. GAME',                 'INCONSISTENT',s32),
    ('33','HOME DOG / GAME 2 OF SERIES',               'INCONSISTENT',s33),
    ('34','HOME FAVORITE AT EXACTLY .500',             'INCONSISTENT',s34),
    ('35','SERIES FACTORS #1',                         'CLEAR FADE',  s35),
    ('36','SERIES FACTOR #2',                          'CLEAR FADE',  s36),
]

# Scenarios needing opponent streak (pass extra arg)
NEEDS_OPP_STREAK = {'09','10','11','12','13'}
NEEDS_OPP_ROAD_WP = {'32'}


# ─────────────────────────────────────────────
# STEP 7: ODDS INPUT FILE HANDLER
# ─────────────────────────────────────────────

def create_odds_template(games, report_date):
    """
    Create a simple Excel template for client to fill in today's odds.
    Columns: Away Team | Home Team | Away Line | Home Line
    """
    fname = ODDS_INPUT
    wb = xlsxwriter.Workbook(fname)
    ws = wb.add_worksheet('Odds Input')

    hdr = wb.add_format({'bold':True,'bg_color':'#2E75B6','font_color':'white',
                          'align':'center','border':1})
    cell = wb.add_format({'align':'center','border':1})
    title = wb.add_format({'bold':True,'font_size':13,'align':'center'})

    ws.merge_range(0,0,0,3, f'DAILY ODDS INPUT — {report_date.strftime("%B %d, %Y")}', title)
    ws.write(1,0,'Away Team',hdr)
    ws.write(1,1,'Home Team',hdr)
    ws.write(1,2,'Away Line (e.g. +130 or -150)',hdr)
    ws.write(1,3,'Home Line (e.g. -130 or +150)',hdr)
    ws.set_column(0,1,28)
    ws.set_column(2,3,28)

    for i, g in enumerate(games):
        ws.write(2+i, 0, g['away_team'], cell)
        ws.write(2+i, 1, g['home_team'], cell)
        ws.write(2+i, 2, '', cell)
        ws.write(2+i, 3, '', cell)

    wb.close()
    print(f'Odds template created: {fname}')
    print(f'Please fill in the moneylines and re-run the script.')

def load_odds(games):
    """Load odds from the filled-in template. Returns dict: team -> line."""
    if not os.path.exists(ODDS_INPUT):
        return {}

    wb = load_workbook(ODDS_INPUT, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    odds = {}
    for row in rows[2:]:
        if row[0] is None: continue
        away_team = str(row[0]).strip().upper()
        home_team = str(row[1]).strip().upper() if row[1] else ''
        away_line = row[2]
        home_line = row[3]
        ln = numeric_line(away_line)
        if ln is not None:
            odds[away_team] = ln
        ln = numeric_line(home_line)
        if ln is not None:
            odds[home_team] = ln
    return odds


# ─────────────────────────────────────────────
# STEP 8: RUN SCENARIOS FOR TODAY'S GAMES
# ─────────────────────────────────────────────

def check_scenarios(game_rows, opp_streaks, opp_road_wpct, opp_lines=None):
    """
    For each team row (away/home), check all 36 scenarios.
    Returns list of triggered scenario results.
    opp_lines: dict of team -> moneyline (used to store correct P/L line for FADE plays)
    """
    if opp_lines is None:
        opp_lines = {}
    triggers = []
    for row in game_rows:
        if row is None: continue
        team    = row['team']
        opp     = row['opponent']
        ha      = row['home_away']
        line    = row['line']

        for sid, sname, verdict, func in SCENARIO_DEFS:
            try:
                if sid in NEEDS_OPP_STREAK:
                    fired = func(row, opp_streaks)
                elif sid in NEEDS_OPP_ROAD_WP:
                    fired = func(row, opp_road_wpct)
                else:
                    fired = func(row)
            except Exception as e:
                print(f'Warning: scenario {sid} error for {team}: {e}')
                fired = False

            if fired:
                if verdict == 'CLEAR FADE':
                    play = f'FADE {team.title()}'
                elif verdict == 'CLEAR BET':
                    play = f'BET {team.title()}'
                else:
                    play = f'WATCH {team.title()}'
                # For FADE plays, P/L is based on betting the OPPONENT (the dog side)
                # Store opponent_line so the Results Tracker uses the correct payout line
                opp_line = opp_lines.get(opp) if verdict == 'CLEAR FADE' else None
                triggers.append({
                    'team':        team,
                    'opponent':    opp,
                    'home_away':   ha,
                    'line':        line,
                    'opp_line':    opp_line,
                    'scenario_id': sid,
                    'scenario':    sname,
                    'verdict':     verdict,
                    'play':        play,
                })
    return triggers


# ─────────────────────────────────────────────
# STEP 9: BUILD EXCEL REPORT
# ─────────────────────────────────────────────

def title_case(name):
    """Convert ALL CAPS team name to Title Case."""
    exceptions = {'Sox','Red','White','Blue','Los','San','New','St.'}
    return ' '.join(w.capitalize() for w in name.split())

def fmt_line(line):
    """Format moneyline for display."""
    if line is None or line == 'N/A': return 'N/A'
    try:
        line = int(line)
    except (TypeError, ValueError):
        return str(line)
    return f'+{line}' if line > 0 else str(line)


def pl_line_for_trigger(t):
    """Moneyline used for P/L — opponent line for FADE, team line otherwise."""
    if t.get('verdict') == 'CLEAR FADE':
        return numeric_line(t.get('opp_line'))
    return numeric_line(t.get('line'))


def numeric_line(line):
    """Coerce moneyline to int for formulas/P/L, or None."""
    if line is None:
        return None
    try:
        v = int(float(line))
        return v
    except (TypeError, ValueError):
        return None


def build_report(games, triggers, report_date, odds):
    """Write CLI daily report to disk. Streamlit downloads use report_builder.build_report_bytes()."""
    fname = os.path.join(OUTPUT_DIR, f'MLB_Daily_Report_{report_date.strftime("%Y-%m-%d")}.xlsx')
    wb = xlsxwriter.Workbook(fname)

    # ── COLOUR PALETTE ───────────────────────────────────────────
    NAVY        = '#1B2A4A'
    STEEL       = '#2E5F8A'
    LIGHT_STEEL = '#D0E4F5'
    FADE_RED    = '#C00000'
    FADE_BG     = '#FFE7E7'
    WATCH_AMB   = '#7D5A00'
    WATCH_BG    = '#FFF3CC'
    AWAY_BG     = '#F0F5FB'
    HOME_BG     = '#FFFFFF'
    GAME_HDR_BG = '#1B2A4A'
    SCEN_BG     = '#E8F0FA'
    GRAY_TEXT   = '#666666'
    GREEN_BG    = '#E2EFDA'
    GREEN_FG    = '#375623'

    # ── FORMATS ──────────────────────────────────────────────────
    # Banner
    f_banner = wb.add_format({
        'bold': True, 'font_size': 16, 'font_name': 'Calibri',
        'font_color': 'white', 'bg_color': NAVY,
        'align': 'center', 'valign': 'vcenter',
    })
    f_subtitle = wb.add_format({
        'font_size': 11, 'font_name': 'Calibri', 'italic': True,
        'font_color': LIGHT_STEEL, 'bg_color': NAVY,
        'align': 'center', 'valign': 'vcenter',
    })
    f_col_hdr = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': 'white', 'bg_color': STEEL,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#1A4A72',
        'text_wrap': False,
    })
    # Game header bar (spans full row, shows "Game N: Away @ Home")
    f_game_hdr = wb.add_format({
        'bold': True, 'font_size': 11, 'font_name': 'Calibri',
        'font_color': 'white', 'bg_color': GAME_HDR_BG,
        'align': 'left', 'valign': 'vcenter',
        'left': 2, 'left_color': STEEL,
        'top': 1, 'top_color': '#0A1828',
        'bottom': 1, 'bottom_color': '#0A1828',
    })
    # Away/Home team label cells
    f_away_label = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': STEEL, 'bg_color': AWAY_BG,
        'align': 'left', 'valign': 'vcenter',
        'left': 2, 'left_color': STEEL,
        'border': 1, 'border_color': '#B8C9DC',
        'indent': 1,
    })
    f_home_label = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': '#333333', 'bg_color': HOME_BG,
        'align': 'left', 'valign': 'vcenter',
        'left': 2, 'left_color': STEEL,
        'border': 1, 'border_color': '#B8C9DC',
        'indent': 1,
    })
    f_away_cell = wb.add_format({
        'font_size': 10, 'font_name': 'Calibri',
        'bg_color': AWAY_BG, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
    })
    f_home_cell = wb.add_format({
        'font_size': 10, 'font_name': 'Calibri',
        'bg_color': HOME_BG, 'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
    })
    # FADE play — red
    f_fade_away = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': FADE_RED, 'bg_color': FADE_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#E8AAAA',
    })
    f_fade_home = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': FADE_RED, 'bg_color': FADE_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#E8AAAA',
    })
    # WATCH play — amber
    f_watch = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': 'Calibri',
        'font_color': WATCH_AMB, 'bg_color': WATCH_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#D4B800',
    })
    # No play — grey dash
    f_no_play = wb.add_format({
        'italic': True, 'font_size': 9, 'font_name': 'Calibri',
        'font_color': GRAY_TEXT, 'bg_color': AWAY_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
    })
    f_no_play_h = wb.add_format({
        'italic': True, 'font_size': 9, 'font_name': 'Calibri',
        'font_color': GRAY_TEXT, 'bg_color': HOME_BG,
        'align': 'center', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
    })
    # Scenario name cell (light blue bg)
    f_scen_away = wb.add_format({
        'font_size': 9, 'font_name': 'Calibri', 'italic': True,
        'font_color': STEEL, 'bg_color': SCEN_BG,
        'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
        'text_wrap': True, 'indent': 1,
    })
    f_scen_home = wb.add_format({
        'font_size': 9, 'font_name': 'Calibri', 'italic': True,
        'font_color': '#555555', 'bg_color': HOME_BG,
        'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC',
        'text_wrap': True, 'indent': 1,
    })
    f_no_scen_away = wb.add_format({
        'italic': True, 'font_size': 9, 'font_color': GRAY_TEXT,
        'bg_color': SCEN_BG, 'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC', 'indent': 1,
    })
    f_no_scen_home = wb.add_format({
        'italic': True, 'font_size': 9, 'font_color': GRAY_TEXT,
        'bg_color': HOME_BG, 'align': 'left', 'valign': 'vcenter',
        'border': 1, 'border_color': '#B8C9DC', 'indent': 1,
    })
    f_no_games = wb.add_format({
        'italic': True, 'font_size': 11, 'font_color': GRAY_TEXT,
        'align': 'center', 'valign': 'vcenter',
    })
    # Summary formats
    f_sum_label  = wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri',
                                   'font_color':NAVY,'bg_color':'#EDF3FB',
                                   'border':1,'border_color':'#B8C9DC',
                                   'align':'left','valign':'vcenter','indent':1})
    f_sum_val    = wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri',
                                   'font_color':NAVY,'bg_color':'#EDF3FB',
                                   'border':1,'border_color':'#B8C9DC',
                                   'align':'center','valign':'vcenter'})
    f_sum_fade_l = wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri',
                                   'font_color':FADE_RED,'bg_color':FADE_BG,
                                   'border':1,'border_color':'#E8AAAA',
                                   'align':'left','valign':'vcenter','indent':1})
    f_sum_fade_v = wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri',
                                   'font_color':FADE_RED,'bg_color':FADE_BG,
                                   'border':1,'border_color':'#E8AAAA',
                                   'align':'center','valign':'vcenter'})
    f_sum_watch_l= wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri',
                                   'font_color':WATCH_AMB,'bg_color':WATCH_BG,
                                   'border':1,'border_color':'#D4B800',
                                   'align':'left','valign':'vcenter','indent':1})
    f_sum_watch_v= wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri',
                                   'font_color':WATCH_AMB,'bg_color':WATCH_BG,
                                   'border':1,'border_color':'#D4B800',
                                   'align':'center','valign':'vcenter'})

    # ── COLUMN LAYOUT ─────────────────────────────────────────────
    # New column layout: GAME | H/A | Team | Odds | Play / Scenario (combined)
    NCOLS = 5
    COL_WIDTHS = [32, 8, 26, 10, 50]
    COL_HEADERS = ['GAME', 'H/A', 'Team', 'Odds', 'Play  /  Scenario']

    # Heavy bottom border formats for home rows (separates games visually)
    f_home_lbl_hb = wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':'#333333','bg_color':HOME_BG,'align':'left','valign':'vcenter','left':2,'left_color':STEEL,'top':1,'top_color':'#B8C9DC','bottom':2,'bottom_color':STEEL,'right':1,'right_color':'#B8C9DC','indent':1})
    f_home_c_hb   = wb.add_format({'font_size':10,'font_name':'Calibri','bg_color':HOME_BG,'align':'center','valign':'vcenter','top':1,'top_color':'#B8C9DC','bottom':2,'bottom_color':STEEL,'left':1,'left_color':'#B8C9DC','right':1,'right_color':'#B8C9DC'})
    f_fade_h_hb   = wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':FADE_RED,'bg_color':FADE_BG,'align':'center','valign':'vcenter','top':1,'top_color':'#E8AAAA','bottom':2,'bottom_color':STEEL,'left':1,'left_color':'#E8AAAA','right':1,'right_color':'#E8AAAA'})
    f_bet_a       = wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':GREEN_FG,'bg_color':GREEN_BG,'align':'center','valign':'vcenter','border':1,'border_color':'#A8D08D'})
    f_bet_h_hb    = wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':GREEN_FG,'bg_color':GREEN_BG,'align':'center','valign':'vcenter','top':1,'bottom':2,'bottom_color':STEEL,'left':1,'right':1,'border_color':'#A8D08D'})
    f_watch_h_hb  = wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':WATCH_AMB,'bg_color':WATCH_BG,'align':'center','valign':'vcenter','top':1,'bottom':2,'bottom_color':STEEL,'left':1,'right':1,'border_color':'#D4B800'})
    f_no_play_h_hb= wb.add_format({'italic':True,'font_size':9,'font_name':'Calibri','font_color':GRAY_TEXT,'bg_color':HOME_BG,'align':'center','valign':'vcenter','top':1,'top_color':'#B8C9DC','bottom':2,'bottom_color':STEEL,'left':1,'left_color':'#B8C9DC','right':1,'right_color':'#B8C9DC'})
    f_scen_h_hb   = wb.add_format({'font_size':9,'font_name':'Calibri','italic':True,'font_color':'#555555','bg_color':HOME_BG,'align':'left','valign':'vcenter','text_wrap':True,'indent':1,'top':1,'top_color':'#B8C9DC','bottom':2,'bottom_color':STEEL,'left':1,'left_color':'#B8C9DC','right':1,'right_color':'#B8C9DC'})
    f_game_cell   = wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri','font_color':NAVY,'bg_color':AWAY_BG,'align':'left','valign':'vcenter','border':1,'border_color':'#B8C9DC','indent':1})
    f_game_cell_h = wb.add_format({'font_size':9,'font_name':'Calibri','font_color':GRAY_TEXT,'bg_color':HOME_BG,'align':'left','valign':'vcenter','top':1,'bottom':2,'bottom_color':STEEL,'left':1,'right':1,'border_color':'#B8C9DC'})

    def get_play_fmt_away(triggers):
        if not triggers: return f_no_play
        v = triggers[0]['verdict']
        if v=='CLEAR FADE': return f_fade_away
        if v=='CLEAR BET':  return f_bet_a
        return f_watch

    def get_play_fmt_home(triggers):
        if not triggers: return f_no_play_h_hb
        v = triggers[0]['verdict']
        if v=='CLEAR FADE': return f_fade_h_hb
        if v=='CLEAR BET':  return f_bet_h_hb
        return f_watch_h_hb

    def write_tab(ws, tab_label, verdict_filter, tab_color):
        ws.set_tab_color(tab_color)
        for ci, w in enumerate(COL_WIDTHS):
            ws.set_column(ci, ci, w)

        ws.set_row(0, 30); ws.set_row(1, 18)
        ws.merge_range(0,0,0,NCOLS-1, f'⚾  MLB DAILY BETTING REPORT  —  {tab_label}', f_banner)
        ws.merge_range(1,0,1,NCOLS-1,
            f'{report_date.strftime("%A, %B %d, %Y")}  |  Generated {datetime.now().strftime("%I:%M %p")}',
            f_subtitle)
        ws.set_row(2, 20)
        for ci, h in enumerate(COL_HEADERS):
            ws.write(2, ci, h, f_col_hdr)
        ws.freeze_panes(3, 0)

        row = 3; games_written = 0

        for g in games:
            away, home = g['away_team'], g['home_team']
            away_tc, home_tc = title_case(away), title_case(home)
            matchup = f'{away_tc}  @  {home_tc}'

            at = [t for t in triggers if t['team']==away and t['opponent']==home and t['verdict']==verdict_filter]
            ht = [t for t in triggers if t['team']==home and t['opponent']==away and t['verdict']==verdict_filter]
            if not at and not ht: continue
            games_written += 1

            # Away row
            ws.set_row(row, 20)
            a_play = (at[0]['play'] + '  |  ' + '  |  '.join(f"#{t['scenario_id']} {t['scenario']}" for t in at)) if at else ''
            ws.write(row, 0, matchup,             f_game_cell)
            ws.write(row, 1, 'AWAY',              f_away_cell)
            ws.write(row, 2, away_tc,             f_away_label)
            ws.write(row, 3, fmt_line(odds.get(away)), f_away_cell)
            ws.write(row, 4, a_play,              get_play_fmt_away(at))
            row += 1

            # Home row — heavy bottom border separates games
            ws.set_row(row, 20)
            h_play = (ht[0]['play'] + '  |  ' + '  |  '.join(f"#{t['scenario_id']} {t['scenario']}" for t in ht)) if ht else ''
            ws.write(row, 0, '',                  f_game_cell_h)
            ws.write(row, 1, 'HOME',              f_home_c_hb)
            ws.write(row, 2, home_tc,             f_home_lbl_hb)
            ws.write(row, 3, fmt_line(odds.get(home)), f_home_c_hb)
            ws.write(row, 4, h_play,              get_play_fmt_home(ht))
            row += 1

        if games_written == 0:
            ws.set_row(row, 30)
            ws.merge_range(row,0,row,NCOLS-1,
                f'No {verdict_filter.title()} scenarios triggered today.', f_no_games)
        return games_written

    # ── WRITE TABS ────────────────────────────────────────────────
    ws_bet  = wb.add_worksheet('🟢 Clear Bet')
    n_bet   = write_tab(ws_bet,  'CLEAR BET',    'CLEAR BET',    '#00B050')

    ws_fade = wb.add_worksheet('🔴 Clear Fade')
    n_fade  = write_tab(ws_fade, 'CLEAR FADE',   'CLEAR FADE',   '#FF0000')

    ws_inc  = wb.add_worksheet('🟡 Inconsistent')
    n_inc   = write_tab(ws_inc,  'INCONSISTENT', 'INCONSISTENT', '#FFD700')

    # ── SUMMARY TAB ───────────────────────────────────────────────
    ws_sum = wb.add_worksheet('📊 Summary')
    ws_sum.set_tab_color('#1B2A4A')
    ws_sum.set_column(0, 0, 35)
    ws_sum.set_column(1, 1, 18)
    ws_sum.set_column(2, 2, 35)
    ws_sum.set_column(3, 3, 18)

    ws_sum.set_row(0, 30)
    ws_sum.set_row(1, 18)
    ws_sum.merge_range(0, 0, 0, 3, '⚾  MLB DAILY BETTING REPORT  —  SUMMARY', f_banner)
    ws_sum.merge_range(1, 0, 1, 3,
        f'{report_date.strftime("%A, %B %d, %Y")}  |  Generated {datetime.now().strftime("%I:%M %p")}',
        f_subtitle)

    n_bet_t  = sum(1 for t in triggers if t['verdict']=='CLEAR BET')
    n_fade_t = sum(1 for t in triggers if t['verdict']=='CLEAR FADE')
    n_inc_t  = sum(1 for t in triggers if t['verdict']=='INCONSISTENT')

    # Add green formats for Clear Bet
    f_sum_bet_l = wb.add_format({'bold':True,'font_size':11,'font_name':'Calibri',
                                  'font_color':GREEN_FG,'bg_color':GREEN_BG,
                                  'border':1,'border_color':'#A8D08D',
                                  'align':'left','valign':'vcenter','indent':1})
    f_sum_bet_v = wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri',
                                  'font_color':GREEN_FG,'bg_color':GREEN_BG,
                                  'border':1,'border_color':'#A8D08D',
                                  'align':'center','valign':'vcenter'})

    ws_sum.set_row(3, 28); ws_sum.set_row(4, 28); ws_sum.set_row(5, 28)
    ws_sum.write(3, 0, 'Total Games Today',       f_sum_label)
    ws_sum.write(3, 1, len(games),                f_sum_val)
    ws_sum.write(3, 2, 'Games With Triggers',     f_sum_label)
    ws_sum.write(3, 3, n_bet + n_fade + n_inc,    f_sum_val)
    ws_sum.write(4, 0, '🟢  Clear Bet Triggers',  f_sum_bet_l)
    ws_sum.write(4, 1, n_bet_t,                   f_sum_bet_v)
    ws_sum.write(4, 2, '🔴  Clear Fade Triggers', f_sum_fade_l)
    ws_sum.write(4, 3, n_fade_t,                  f_sum_fade_v)
    ws_sum.write(5, 0, '🟡  Inconsistent Triggers',f_sum_watch_l)
    ws_sum.write(5, 1, n_inc_t,                   f_sum_watch_v)

    ws_sum.set_row(7, 20)
    ws_sum.merge_range(6, 0, 6, 3, 'TOP PLAYS TODAY', f_col_hdr)
    ws_sum.write(7, 0, 'Team', f_col_hdr)
    ws_sum.write(7, 1, 'Odds', f_col_hdr)
    ws_sum.write(7, 2, 'Play', f_col_hdr)
    ws_sum.write(7, 3, 'Scenario', f_col_hdr)

    sum_row = 8
    for t in [x for x in triggers if x['verdict']=='CLEAR BET']:
        ws_sum.write(sum_row, 0, title_case(t['team']), f_sum_bet_l)
        ws_sum.write(sum_row, 1, fmt_line(t['line']),   f_sum_bet_v)
        ws_sum.write(sum_row, 2, t['play'],             f_sum_bet_l)
        ws_sum.write(sum_row, 3, f"#{t['scenario_id']} {t['scenario']}", f_sum_bet_l)
        sum_row += 1
    for t in [x for x in triggers if x['verdict']=='CLEAR FADE']:
        ws_sum.write(sum_row, 0, title_case(t['team']), f_sum_fade_l)
        ws_sum.write(sum_row, 1, fmt_line(t['line']),   f_sum_fade_v)
        ws_sum.write(sum_row, 2, t['play'],             f_sum_fade_l)
        ws_sum.write(sum_row, 3, f"#{t['scenario_id']} {t['scenario']}", f_sum_fade_l)
        sum_row += 1
    for t in [x for x in triggers if x['verdict']=='INCONSISTENT']:
        ws_sum.write(sum_row, 0, title_case(t['team']), f_sum_watch_l)
        ws_sum.write(sum_row, 1, fmt_line(t['line']),   f_sum_watch_v)
        ws_sum.write(sum_row, 2, t['play'],             f_sum_watch_l)
        ws_sum.write(sum_row, 3, f"#{t['scenario_id']} {t['scenario']}", f_sum_watch_l)
        sum_row += 1

    if sum_row == 8:
        ws_sum.merge_range(8, 0, 8, 3, 'No scenarios triggered today.', f_no_games)
    ws_sum.freeze_panes(8, 0)

    # ── RESULTS TRACKER TAB ───────────────────────────────────────
    # Client enters W/L results here; tab auto-tallies by scenario
    ws_track = wb.add_worksheet('📈 Results Tracker')
    ws_track.set_tab_color('#375623')
    ws_track.set_column(0, 0, 12)   # Date
    ws_track.set_column(1, 1, 26)   # Team
    ws_track.set_column(2, 2, 10)   # H/A
    ws_track.set_column(3, 3, 10)   # Odds
    ws_track.set_column(4, 4, 14)   # Play
    ws_track.set_column(5, 5, 40)   # Scenario
    ws_track.set_column(6, 6, 12)   # Classification
    ws_track.set_column(7, 7, 12)   # Result (W/L — client fills)
    ws_track.set_column(8, 8, 14)   # Net P/L (auto)
    ws_track.set_column(9, 9, None, None, {'hidden': True})  # PayoutLine
    ws_track.set_column(10, 10, None, None, {'hidden': True})  # Opponent

    f_track_hdr = wb.add_format({'bold':True,'font_color':'white','bg_color':'#375623',
                                   'align':'center','valign':'vcenter','border':1,'font_name':'Calibri'})
    f_track_note= wb.add_format({'italic':True,'font_color':GRAY_TEXT,'font_size':9,
                                   'align':'center','valign':'vcenter','border':1})
    f_track_cell= wb.add_format({'align':'center','valign':'vcenter','border':1,
                                   'border_color':'#C6EFCE','font_name':'Calibri'})
    f_track_left= wb.add_format({'align':'left','valign':'vcenter','border':1,
                                   'border_color':'#C6EFCE','font_name':'Calibri','indent':1})
    f_input_cell= wb.add_format({'bold':True,'align':'center','valign':'vcenter',
                                   'border':2,'border_color':'#375623',
                                   'bg_color':'#EAF4E8','font_name':'Calibri'})

    ws_track.set_row(0, 30); ws_track.set_row(1, 18)
    ws_track.merge_range(0, 0, 0, 8, '📈  RESULTS TRACKER  —  Enter W or L after each game', f_banner)
    ws_track.merge_range(1, 0, 1, 8,
        'Enter W or L in the Result column. Net P/L calculates automatically.',
        f_subtitle)

    track_headers = ['Date', 'Team', 'H/A', 'Odds', 'Play', 'Scenario', 'Type', 'Result\n(W or L)', 'Net P/L\n($100 flat)', 'PayoutLine', 'Opponent']
    ws_track.set_row(2, 30)
    for ci, h in enumerate(track_headers):
        ws_track.write(2, ci, h, f_track_hdr)

    track_row = 3
    for t in triggers:
        line = t['line']
        pl_line = numeric_line(pl_line_for_trigger(t))
        excel_row = track_row + 1
        ws_track.write(track_row, 0, report_date.strftime('%Y-%m-%d'), f_track_cell)
        ws_track.write(track_row, 1, title_case(t['team']),             f_track_left)
        ws_track.write(track_row, 2, t['home_away'].upper(),            f_track_cell)
        ws_track.write(track_row, 3, fmt_line(line),                    f_track_cell)
        ws_track.write(track_row, 4, t['play'],                         f_track_left)
        ws_track.write(track_row, 5, f"#{t['scenario_id']} {t['scenario']}", f_track_left)
        ws_track.write(track_row, 6, t['verdict'],                      f_track_cell)
        ws_track.write(track_row, 7, '',                                f_input_cell)
        if pl_line is not None:
            if pl_line > 0:
                payout_formula = f'=IF(H{excel_row}="W",{pl_line},IF(H{excel_row}="L",-100,""))'
            else:
                payout_formula = f'=IF(H{excel_row}="W",ROUND(100/ABS({pl_line})*100,2),IF(H{excel_row}="L",-100,""))'
            ws_track.write_formula(track_row, 8, payout_formula, f_track_cell)
        else:
            ws_track.write(track_row, 8, '', f_track_cell)
        if pl_line is not None:
            ws_track.write(track_row, 9, pl_line, f_track_cell)
        ws_track.write(track_row, 10, str(t.get('opponent', '') or '').strip().upper(), f_track_cell)
        track_row += 1

    # Totals row
    if track_row > 3:
        f_total = wb.add_format({'bold':True,'bg_color':'#375623','font_color':'white',
                                   'align':'center','border':1,'font_name':'Calibri'})
        ws_track.merge_range(track_row, 0, track_row, 7, 'TOTAL NET P/L', f_total)
        ws_track.write_formula(track_row, 8,
            f'=SUMIF(H4:H{track_row},"W",I4:I{track_row})+SUMIF(H4:H{track_row},"L",I4:I{track_row})',
            f_total)

    ws_track.freeze_panes(3, 0)
    ws_track.autofilter(2, 0, track_row, 8)

    # ── CUMULATIVE RESULTS (prior days from Master_Results.xlsx) ───
    cum_data_end = 3
    hist_rows = load_master_history_before(report_date)
    has_cumulative = len(hist_rows) > 0
    if has_cumulative:
        w_cum = wb.add_worksheet('Cumulative Results')
        w_cum.hide()
        w_cum.set_column(0, 0, 12)
        w_cum.set_column(1, 1, 26)
        w_cum.set_column(5, 5, 40)
        w_cum.set_column(7, 7, 14)
        w_cum.set_column(8, 8, 16)
        w_cum.set_column(9, 9, None, None, {'hidden': True})
        w_cum.set_column(10, 10, None, None, {'hidden': True})
        cum_headers = ['Date', 'Team', 'H/A', 'Odds', 'Play', 'Scenario', 'Type',
                       'Result (W/L)', 'Net P/L ($100)', 'PayoutLine', 'Opponent']
        for ci, h in enumerate(cum_headers):
            w_cum.write(2, ci, h, f_track_hdr)
        cum_row = 3
        for row in hist_rows:
            er = cum_row + 1
            w_cum.write(cum_row, 0, str(row.get('Date', '')), f_track_cell)
            w_cum.write(cum_row, 1, str(row.get('Team', '')), f_track_left)
            w_cum.write(cum_row, 2, str(row.get('H/A', '')), f_track_cell)
            w_cum.write(cum_row, 3, str(row.get('Odds', '')), f_track_cell)
            w_cum.write(cum_row, 4, str(row.get('Play', '')), f_track_left)
            w_cum.write(cum_row, 5, str(row.get('Scenario', '')), f_track_left)
            w_cum.write(cum_row, 6, str(row.get('Type', '')), f_track_cell)
            res = str(row.get('Result', '') or '').strip().upper()
            w_cum.write(cum_row, 7, res if res in ('W', 'L') else '', f_input_cell if res in ('W', 'L') else f_track_cell)
            payout = numeric_line(row.get('PayoutLine'))
            if payout is None:
                payout = numeric_line(str(row.get('Odds', '')).replace('+', ''))
            if payout is not None:
                pf = (f'=IF(H{er}="W",{payout},IF(H{er}="L",-100,""))' if payout > 0
                      else f'=IF(H{er}="W",ROUND(100/ABS({payout})*100,2),IF(H{er}="L",-100,""))')
                w_cum.write_formula(cum_row, 8, pf, f_track_cell)
                w_cum.write(cum_row, 9, payout, f_track_cell)
            else:
                w_cum.write(cum_row, 8, '', f_track_cell)
            w_cum.write(cum_row, 10, str(row.get('Opponent', row.get('_opponent', '')) or '').strip().upper(), f_track_cell)
            cum_row += 1
        cum_data_end = cum_row

    # ── SCENARIO PERFORMANCE TAB ──────────────────────────────────
    # Lists all 36 scenarios with cumulative W/L/P/L fed from Results Tracker
    ws_perf = wb.add_worksheet('📋 Scenario Performance')
    ws_perf.set_tab_color('#1F3864')
    ws_perf.set_column(0, 0, 6)    # Scenario #
    ws_perf.set_column(1, 1, 40)   # Scenario Name
    ws_perf.set_column(2, 2, 14)   # Classification
    ws_perf.set_column(3, 3, 8)    # W
    ws_perf.set_column(4, 4, 8)    # L
    ws_perf.set_column(5, 5, 10)   # Total
    ws_perf.set_column(6, 6, 10)   # Win%
    ws_perf.set_column(7, 7, 14)   # Net P/L

    f_perf_banner = wb.add_format({'bold':True,'font_size':14,'font_name':'Calibri',
                                    'font_color':'white','bg_color':'#1F3864',
                                    'align':'center','valign':'vcenter'})
    f_perf_sub    = wb.add_format({'font_size':10,'font_name':'Calibri','italic':True,
                                    'font_color':'#D0E4F5','bg_color':'#1F3864',
                                    'align':'center','valign':'vcenter'})
    f_perf_hdr    = wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri',
                                    'font_color':'white','bg_color':'#2E5F8A',
                                    'align':'center','valign':'vcenter','border':1})
    f_perf_sid    = wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri',
                                    'font_color':'#1F3864','bg_color':'#EDF3FB',
                                    'align':'center','valign':'vcenter','border':1})
    f_perf_name   = wb.add_format({'font_size':10,'font_name':'Calibri',
                                    'font_color':'#1F3864','bg_color':'#EDF3FB',
                                    'align':'left','valign':'vcenter','border':1,'indent':1})
    f_perf_bet    = wb.add_format({'font_size':9,'font_name':'Calibri','bold':True,
                                    'font_color':'#375623','bg_color':'#E2EFDA',
                                    'align':'center','valign':'vcenter','border':1})
    f_perf_fade   = wb.add_format({'font_size':9,'font_name':'Calibri','bold':True,
                                    'font_color':'#C00000','bg_color':'#FFE7E7',
                                    'align':'center','valign':'vcenter','border':1})
    f_perf_inc    = wb.add_format({'font_size':9,'font_name':'Calibri','bold':True,
                                    'font_color':'#7D5A00','bg_color':'#FFF3CC',
                                    'align':'center','valign':'vcenter','border':1})
    f_perf_num    = wb.add_format({'font_size':10,'font_name':'Calibri',
                                    'bg_color':'#EDF3FB','align':'center',
                                    'valign':'vcenter','border':1})
    f_perf_pct    = wb.add_format({'font_size':10,'font_name':'Calibri','num_format':'0.000',
                                    'bg_color':'#EDF3FB','align':'center',
                                    'valign':'vcenter','border':1})
    f_perf_money  = wb.add_format({'font_size':10,'font_name':'Calibri',
                                    'num_format':'$#,##0.00','bg_color':'#EDF3FB',
                                    'align':'center','valign':'vcenter','border':1})
    f_perf_zero   = wb.add_format({'font_size':10,'font_name':'Calibri','italic':True,
                                    'font_color':'#AAAAAA','bg_color':'#EDF3FB',
                                    'align':'center','valign':'vcenter','border':1})
    f_perf_total  = wb.add_format({'bold':True,'font_size':10,'font_name':'Calibri',
                                    'font_color':'white','bg_color':'#1F3864',
                                    'align':'center','valign':'vcenter','border':1})

    # Banner
    ws_perf.set_row(0, 28); ws_perf.set_row(1, 16); ws_perf.set_row(2, 22)
    tr_end = max(track_row + 500, 1000)
    tr_sc = f"'📈 Results Tracker'!F$4:F${tr_end}"
    tr_rc = f"'📈 Results Tracker'!H$4:H${tr_end}"
    tr_pc = f"'📈 Results Tracker'!I$4:I${tr_end}"
    if has_cumulative:
        ws_perf.merge_range(0, 0, 0, 7, '📋  SCENARIO PERFORMANCE TRACKER  —  Season Cumulative', f_perf_banner)
        ws_perf.merge_range(1, 0, 1, 7,
            'Prior days from Master_Results.xlsx + today from Results Tracker (enter W/L there).',
            f_perf_sub)
        ws_perf.merge_range(2, 0, 2, 7,
            'Keep Master_Results.xlsx updated in this folder for season-long CLI tracking.',
            f_perf_sub)
        cum_end = max(cum_data_end + 500, 1000)
        cum_sc = f"'Cumulative Results'!F$4:F${cum_end}"
        cum_rc = f"'Cumulative Results'!H$4:H${cum_end}"
        cum_pc = f"'Cumulative Results'!I$4:I${cum_end}"
    else:
        ws_perf.merge_range(0, 0, 0, 7, '📋  SCENARIO PERFORMANCE TRACKER  —  Today\'s Results Only', f_perf_banner)
        ws_perf.merge_range(1, 0, 1, 7,
            'Updates automatically as you enter results in the Results Tracker tab.', f_perf_sub)
        ws_perf.merge_range(2, 0, 2, 7,
            'Place Master_Results.xlsx in this folder to include prior days here.', f_perf_sub)

    # Column headers
    ws_perf.set_row(3, 20)
    for ci, h in enumerate(['#', 'Scenario Name', 'Classification', 'W', 'L', 'Total', 'Win%', 'Net P/L']):
        ws_perf.write(3, ci, h, f_perf_hdr)
    ws_perf.freeze_panes(4, 0)

    perf_row = 4
    for sid, sname, verdict, _ in SCENARIO_DEFS:
        scen_id_str = f'#{sid} {sname}'

        if verdict == 'CLEAR BET':   vfmt = f_perf_bet;  vlabel = 'CLEAR BET'
        elif verdict == 'CLEAR FADE': vfmt = f_perf_fade; vlabel = 'CLEAR FADE'
        else:                         vfmt = f_perf_inc;  vlabel = 'INCONSISTENT'

        er = perf_row + 1

        ws_perf.write(perf_row, 0, sid,     f_perf_sid)
        ws_perf.write(perf_row, 1, sname,   f_perf_name)
        ws_perf.write(perf_row, 2, vlabel,  vfmt)

        if has_cumulative:
            ws_perf.write_formula(perf_row, 3,
                f'=IFERROR(COUNTIFS({cum_sc},"{scen_id_str}",{cum_rc},"W"),0)+IFERROR(COUNTIFS({tr_sc},"{scen_id_str}",{tr_rc},"W"),0)',
                f_perf_num)
            ws_perf.write_formula(perf_row, 4,
                f'=IFERROR(COUNTIFS({cum_sc},"{scen_id_str}",{cum_rc},"L"),0)+IFERROR(COUNTIFS({tr_sc},"{scen_id_str}",{tr_rc},"L"),0)',
                f_perf_num)
            ws_perf.write_formula(perf_row, 7,
                f'=IFERROR(SUMIFS({cum_pc},{cum_sc},"{scen_id_str}"),0)+IFERROR(SUMIFS({tr_pc},{tr_sc},"{scen_id_str}"),0)',
                f_perf_money)
        else:
            ws_perf.write_formula(perf_row, 3,
                f'=IFERROR(COUNTIFS({tr_sc},"{scen_id_str}",{tr_rc},"W"),0)', f_perf_num)
            ws_perf.write_formula(perf_row, 4,
                f'=IFERROR(COUNTIFS({tr_sc},"{scen_id_str}",{tr_rc},"L"),0)', f_perf_num)
            ws_perf.write_formula(perf_row, 7,
                f'=IFERROR(SUMIFS({tr_pc},{tr_sc},"{scen_id_str}"),0)', f_perf_money)

        ws_perf.write_formula(perf_row, 5, f'=D{er}+E{er}', f_perf_num)
        ws_perf.write_formula(perf_row, 6, f'=IF(F{er}>0,D{er}/F{er},"")', f_perf_pct)

        perf_row += 1

    # Totals row
    ws_perf.set_row(perf_row, 20)
    ws_perf.write(perf_row, 0, '', f_perf_total)
    ws_perf.write(perf_row, 1, 'SEASON TOTALS' if has_cumulative else 'TODAY TOTALS', f_perf_total)
    ws_perf.write(perf_row, 2, '', f_perf_total)
    ws_perf.write_formula(perf_row, 3, f'=SUM(D5:D{perf_row})', f_perf_total)
    ws_perf.write_formula(perf_row, 4, f'=SUM(E5:E{perf_row})', f_perf_total)
    ws_perf.write_formula(perf_row, 5, f'=SUM(F5:F{perf_row})', f_perf_total)
    ws_perf.write_formula(perf_row, 6,
        f'=IF(F{perf_row+1}>0,D{perf_row+1}/F{perf_row+1},"")', f_perf_total)
    ws_perf.write_formula(perf_row, 7, f'=SUM(H5:H{perf_row})', f_perf_total)

    wb.close()
    print(f'\n✓ Report saved: {fname}')
    print(f'  Clear Bet games: {n_bet} ({n_bet_t} triggers)')
    print(f'  Clear Fade games: {n_fade} ({n_fade_t} triggers)')
    print(f'  Inconsistent games: {n_inc} ({n_inc_t} triggers)')
    return fname


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

if __name__ == '__main__':
    print(f'=== MLB Daily Report — {REPORT_DATE} ===\n')

    # 1. Load historical data
    print('Loading historical data...')
    df = load_historical_data()

    # 2. Fetch recent results to bring data current
    print('Checking for recent game results...')
    df, api_warnings = fetch_recent_results(df, REPORT_DATE)
    for msg in api_warnings:
        print(f'Warning: {msg}')

    # 3. Compute team states
    print('Computing team states...')
    enriched = compute_all_states(df)

    # 4. Fetch today's schedule
    print('Fetching today\'s schedule...')
    games = fetch_todays_schedule(REPORT_DATE)
    if not games:
        print('No games today. Exiting.')
        sys.exit(0)

    # 5. Load odds — create template if missing
    odds = load_odds(games)
    if odds:
        save_odds_to_cache(REPORT_DATE, odds)
    if not odds:
        print('\nNo odds found. Creating odds input template...')
        create_odds_template(games, REPORT_DATE)
        print('\nFill in the moneylines in daily_odds_input.xlsx then re-run.')
        sys.exit(0)

    # 6. Build opponent streak lookup for today
    opp_streaks = {}
    opp_road_wpct = {}
    for team in API_TO_CANONICAL.values():
        tdf = enriched[(enriched['team']==team) & (enriched['date'].dt.date < REPORT_DATE)]
        if not tdf.empty:
            last = tdf.iloc[-1]
            sb = last['streak_before']
            res = last['result']
            opp_streaks[team] = (sb+1 if sb>=0 else 1) if res=='W' else (sb-1 if sb<=0 else -1)
            # Road win%
            road = tdf[tdf['home_away']=='away']
            if not road.empty:
                rw = (road['result']=='W').sum()
                rl = (road['result']=='L').sum()
                opp_road_wpct[team] = rw/(rw+rl) if (rw+rl)>0 else None

    # 7. Build game rows and check scenarios
    print('Checking scenarios...')
    all_triggers = []
    for g in games:
        away, home = g['away_team'], g['home_team']
        away_line = odds.get(away)
        home_line = odds.get(home)

        away_state = get_team_state(enriched, away, REPORT_DATE)
        home_state = get_team_state(enriched, home, REPORT_DATE)

        away_row = build_game_row(away_state, 'away', home, away_line)
        home_row = build_game_row(home_state, 'home', away, home_line)

        triggers = check_scenarios([away_row, home_row], opp_streaks, opp_road_wpct, odds)
        all_triggers.extend(triggers)

    # 8. Build report
    build_report(games, all_triggers, REPORT_DATE, odds)
