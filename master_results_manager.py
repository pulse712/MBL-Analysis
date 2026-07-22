"""
Master Results Manager
======================
Manages the cumulative Master_Results.xlsx file that stores all betting results
across the entire season for scenario performance tracking.
"""

import io
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime

from daily_report import pl_line_for_trigger, numeric_line


MASTER_FILE = 'Master_Results.xlsx'
RESULTS_COLUMNS = ['Date', 'Team', 'H/A', 'Odds', 'Play', 'Scenario', 'Type', 'Result', 'Net P/L']


def _detect_header_row(probe_df):
    """Return the row index that contains the 'Date' column header."""
    for i in range(min(6, len(probe_df))):
        row_vals = [str(v).strip() for v in probe_df.iloc[i].tolist()]
        if 'Date' in row_vals:
            return i
    return 2


def _find_results_sheet(sheet_names):
    """
    Pick the sheet that holds trackable results.
    Accepts Master_Results.xlsx and daily reports (Results Tracker tab).
    """
    if 'Master Results' in sheet_names:
        return 'Master Results', 'Master_Results.xlsx'
    for name in sheet_names:
        if name == 'Results Tracker' or 'Results Tracker' in name:
            return name, 'daily report'
    return None, None


def _normalize_results_df(df, raw_ncols=None):
    """Normalize uploaded results to the standard layout (incl. hidden PayoutLine + Opponent)."""
    if raw_ncols is None:
        raw_ncols = len(df.columns)
    ncols = len(df.columns)
    if ncols >= 11:
        df = df.iloc[:, :11].copy()
        df.columns = RESULTS_COLUMNS + ['PayoutLine', 'Opponent']
    elif ncols >= 10:
        df = df.iloc[:, :10].copy()
        df.columns = RESULTS_COLUMNS + ['PayoutLine']
        df['Opponent'] = ''
    else:
        df = df.iloc[:, :9].copy()
        df.columns = RESULTS_COLUMNS
        df['PayoutLine'] = None
        df['Opponent'] = ''
    df = df[df['Date'].notna()]
    # Normalize Date to YYYY-MM-DD string regardless of how pandas read it
    def _fmt_date(v):
        try:
            return pd.to_datetime(v).strftime('%Y-%m-%d')
        except Exception:
            return str(v).strip()
    df['Date'] = df['Date'].apply(_fmt_date)
    date_str = df['Date'].astype(str).str.strip().str.upper()
    team_str = df['Team'].astype(str).str.strip().str.upper()
    mask = (
        ~date_str.str.contains('TOTAL', na=False)
        & ~team_str.str.contains('TOTAL', na=False)
        & date_str.str.match(r'\d{4}-\d{2}-\d{2}', na=False)
    )
    df = df.loc[mask].copy()
    df['Result'] = df['Result'].astype(str).str.strip().str.upper()
    df.loc[~df['Result'].isin(['W', 'L']), 'Result'] = ''
    if 'PayoutLine' in df.columns:
        def _to_int_line(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return None
            try:
                return int(float(v))
            except (TypeError, ValueError):
                return None
        df['PayoutLine'] = df['PayoutLine'].apply(_to_int_line)
    if 'Opponent' in df.columns:
        df['Opponent'] = df['Opponent'].astype(str).str.strip().str.upper()
        df.loc[df['Opponent'].isin(['', 'NAN', 'NONE']), 'Opponent'] = ''
    return df.reset_index(drop=True)


def parse_results_upload(file_bytes):
    """
    Load trackable results from an uploaded Excel file.

    Accepts:
    - Master_Results.xlsx (Master Results sheet)
    - MLB_Daily_Report_*.xlsx (Results Tracker sheet, with W/L entered)

    Returns:
        (DataFrame, source_label, sheet_name, upload_notes)
    """
    xl = pd.ExcelFile(io.BytesIO(file_bytes))
    sheet, source = _find_results_sheet(xl.sheet_names)

    if sheet is None:
        for name in xl.sheet_names:
            probe = pd.read_excel(io.BytesIO(file_bytes), sheet_name=name, header=None, nrows=6)
            if 'Date' in [str(v).strip() for v in probe.iloc[_detect_header_row(probe)].tolist()]:
                sheet, source = name, f'sheet "{name}"'
                break

    if sheet is None:
        raise ValueError(
            "No results data found. Upload Master_Results.xlsx (recommended) or a saved "
            "daily report whose Results Tracker tab has W/L results entered."
        )

    probe = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, header=None, nrows=6)
    skip = _detect_header_row(probe)
    raw_probe = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, skiprows=skip, nrows=1)
    raw_ncols = len(raw_probe.columns)
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, skiprows=skip)
    df = _normalize_results_df(df, raw_ncols=raw_ncols)

    upload_notes = []
    if raw_ncols < 10:
        upload_notes.append(
            'Legacy file missing hidden PayoutLine column (J). '
            'Re-download Master_Results.xlsx from the app for accurate FADE P/L.'
        )
    if raw_ncols < 11:
        upload_notes.append(
            'Legacy file missing hidden Opponent column (K). '
            'Re-download from the app for doubleheader-safe deduplication.'
        )

    if df.empty:
        raise ValueError(
            f"Found sheet '{sheet}' but no result rows. "
            "Open the daily report, enter W/L on the Results Tracker tab, save, then upload."
        )

    return df, source, sheet, upload_notes


def initialize_master_results():
    """
    Create a new Master_Results.xlsx file if it doesn't exist.
    This file has the same structure as the Results Tracker tab.
    """
    if os.path.exists(MASTER_FILE):
        print(f'Master results file already exists: {MASTER_FILE}')
        return

    # Create a new workbook with the Results Tracker structure
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Master Results'

    # Define styles
    header_fill = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', name='Calibri')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set column widths
    ws.column_dimensions['A'].width = 12   # Date
    ws.column_dimensions['B'].width = 26   # Team
    ws.column_dimensions['C'].width = 10   # H/A
    ws.column_dimensions['D'].width = 10   # Odds
    ws.column_dimensions['E'].width = 14   # Play
    ws.column_dimensions['F'].width = 40   # Scenario
    ws.column_dimensions['G'].width = 14   # Type
    ws.column_dimensions['H'].width = 14   # Result
    ws.column_dimensions['I'].width = 16   # Net P/L
    ws.column_dimensions['J'].width = 0.1  # PayoutLine (hidden helper)
    ws.column_dimensions['J'].hidden = True
    ws.column_dimensions['K'].width = 0.1  # Opponent (hidden helper)
    ws.column_dimensions['K'].hidden = True

    # Title row
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = 'MASTER RESULTS TRACKER — All Season Results'
    title_cell.font = Font(bold=True, size=16, color='FFFFFF', name='Calibri')
    title_cell.fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Subtitle row
    ws.merge_cells('A2:I2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = 'Copy results from daily reports into this file to build cumulative season statistics'
    subtitle_cell.font = Font(italic=True, size=10, color='D0E4F5', name='Calibri')
    subtitle_cell.fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # Header row
    headers = ['Date', 'Team', 'H/A', 'Odds', 'Play', 'Scenario', 'Type', 'Result (W/L)', 'Net P/L ($100)', 'PayoutLine', 'Opponent']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    ws.row_dimensions[3].height = 30

    # Add data validation for Result column (H)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"W,L"', allow_blank=True)
    dv.error = 'Please enter W or L'
    dv.errorTitle = 'Invalid Entry'
    ws.add_data_validation(dv)
    dv.add('H4:H10000')

    # Freeze panes
    ws.freeze_panes = 'A4'

    wb.save(MASTER_FILE)
    print(f'✓ Created new master results file: {MASTER_FILE}')


def append_to_master_results(triggers, report_date):
    """
    Append today's triggers to the Master_Results.xlsx file.
    This allows the client to track results across multiple days.
    
    Args:
        triggers: List of trigger dictionaries from today's report
        report_date: Date object for today's report
    """
    if not os.path.exists(MASTER_FILE):
        initialize_master_results()

    # Load existing data
    wb = load_workbook(MASTER_FILE)
    ws = wb.active
    ws.column_dimensions['J'].width = 0.1
    ws.column_dimensions['J'].hidden = True
    ws.column_dimensions['K'].width = 0.1
    ws.column_dimensions['K'].hidden = True
    if ws.cell(row=3, column=10).value != 'PayoutLine':
        ws.cell(row=3, column=10).value = 'PayoutLine'
    if ws.cell(row=3, column=11).value != 'Opponent':
        ws.cell(row=3, column=11).value = 'Opponent'

    # Remove prior rows for this date (re-generate replaces stale triggers)
    date_str = report_date.strftime('%Y-%m-%d')
    r = 4
    while ws.cell(row=r, column=1).value is not None:
        if str(ws.cell(row=r, column=1).value)[:10] == date_str:
            ws.delete_rows(r, 1)
        else:
            r += 1

    # Find the next empty row
    next_row = 4  # Start after header (row 3)
    while ws.cell(row=next_row, column=1).value is not None:
        next_row += 1

    # Define cell styles
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    cell_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center', indent=1)
    thin_border = Border(
        left=Side(style='thin', color='C6EFCE'),
        right=Side(style='thin', color='C6EFCE'),
        top=Side(style='thin', color='C6EFCE'),
        bottom=Side(style='thin', color='C6EFCE')
    )
    input_fill = PatternFill(start_color='EAF4E8', end_color='EAF4E8', fill_type='solid')
    input_font = Font(bold=True, name='Calibri')
    input_border = Border(
        left=Side(style='medium', color='375623'),
        right=Side(style='medium', color='375623'),
        top=Side(style='medium', color='375623'),
        bottom=Side(style='medium', color='375623')
    )

    # Add each trigger as a row
    for trigger in triggers:
        row = next_row
        
        # Date
        cell = ws.cell(row=row, column=1)
        cell.value = report_date.strftime('%Y-%m-%d')
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Team
        cell = ws.cell(row=row, column=2)
        cell.value = trigger['team'].title()
        cell.alignment = left_alignment
        cell.border = thin_border

        # H/A
        cell = ws.cell(row=row, column=3)
        cell.value = trigger['home_away'].upper()
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Odds
        cell = ws.cell(row=row, column=4)
        line = trigger['line']
        if line is not None:
            cell.value = f'+{line}' if line > 0 else str(line)
        else:
            cell.value = 'N/A'
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Play
        cell = ws.cell(row=row, column=5)
        cell.value = trigger['play']
        cell.alignment = left_alignment
        cell.border = thin_border

        # Scenario
        cell = ws.cell(row=row, column=6)
        cell.value = f"#{trigger['scenario_id']} {trigger['scenario']}"
        cell.alignment = left_alignment
        cell.border = thin_border

        # Type
        cell = ws.cell(row=row, column=7)
        cell.value = trigger['verdict']
        cell.alignment = cell_alignment
        cell.border = thin_border

        # Result (W/L) - empty for client to fill
        cell = ws.cell(row=row, column=8)
        cell.value = ''
        cell.alignment = cell_alignment
        cell.fill = input_fill
        cell.font = input_font
        cell.border = input_border

        # Net P/L - formula (uses payout line so FADE bets pay on opponent odds)
        cell = ws.cell(row=row, column=9)
        payout_line = numeric_line(pl_line_for_trigger(trigger))
        if payout_line is not None:
            if payout_line > 0:
                formula = f'=IF(H{row}="W",{payout_line},IF(H{row}="L",-100,""))'
            else:
                formula = f'=IF(H{row}="W",ROUND(100/ABS({payout_line})*100,2),IF(H{row}="L",-100,""))'
            cell.value = formula
        else:
            cell.value = ''
        cell.alignment = cell_alignment
        cell.border = thin_border

        if payout_line is not None:
            pc = ws.cell(row=row, column=10)
            pc.value = payout_line
            pc.alignment = cell_alignment
            pc.border = thin_border

        opp_cell = ws.cell(row=row, column=11)
        opp_cell.value = str(trigger.get('opponent', '') or '').strip().upper()
        opp_cell.alignment = cell_alignment
        opp_cell.border = thin_border

        next_row += 1

    wb.save(MASTER_FILE)
    print(f'✓ Appended {len(triggers)} triggers to {MASTER_FILE}')


def load_master_results():
    """
    Load all results from Master_Results.xlsx and return as a DataFrame.
    Returns None if file doesn't exist or is empty.
    """
    if not os.path.exists(MASTER_FILE):
        return None

    try:
        with open(MASTER_FILE, 'rb') as f:
            df, _, _, _ = parse_results_upload(f.read())
        return None if df.empty else df
    except Exception as e:
        print(f'Warning: Could not load master results: {e}')
        return None


def _payout_line_from_row(row):
    """Resolve the moneyline used for P/L from a results row."""
    if 'PayoutLine' in row.index and pd.notna(row.get('PayoutLine')):
        pl = numeric_line(row.get('PayoutLine'))
        if pl is not None:
            return pl
    verdict = str(row.get('Type', '')).strip().upper()
    if verdict == 'CLEAR FADE':
        return None
    odds_val = str(row.get('Odds', '')).strip().replace(' ', '')
    if odds_val.upper() in ('', 'N/A', 'NAN'):
        return None
    return numeric_line(odds_val.replace('+', ''))


def _net_pl_from_result(result, payout_line):
    """Calculate net P/L in dollars (based on $100 bet) from result + payout line."""
    if result not in ('W', 'L') or payout_line is None:
        return 0.0
    if result == 'L':
        return -100.0
    if payout_line > 0:
        return float(payout_line)
    return round(100 / abs(payout_line) * 100, 2)


def get_cumulative_stats():
    """
    Calculate cumulative statistics from Master_Results.xlsx.
    Returns a dictionary with stats for each scenario.
    """
    df = load_master_results()
    if df is None or df.empty:
        return {}

    stats = {}
    
    # Group by scenario
    for scenario in df['Scenario'].unique():
        scenario_df = df[df['Scenario'] == scenario]
        
        # Count wins and losses
        wins = (scenario_df['Result'] == 'W').sum()
        losses = (scenario_df['Result'] == 'L').sum()
        total = wins + losses
        win_pct = wins / total if total > 0 else 0
        
        # Sum net P/L from Result + payout line (Excel formulas are not readable via pandas)
        pl_values = scenario_df.apply(
            lambda r: _net_pl_from_result(r['Result'], _payout_line_from_row(r)),
            axis=1,
        )
        net_pl = pl_values.sum()
        
        stats[scenario] = {
            'wins': wins,
            'losses': losses,
            'total': total,
            'win_pct': win_pct,
            'net_pl': net_pl
        }
    
    return stats


def get_master_results_path():
    """Return the full path to the master results file."""
    return os.path.abspath(MASTER_FILE)


if __name__ == '__main__':
    # Test: Initialize master results file
    initialize_master_results()
    print(f'\nMaster results file location: {get_master_results_path()}')
