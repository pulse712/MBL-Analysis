"""
MLB Daily Betting Report - Streamlit Web App
"""
import streamlit as st
import pandas as pd
import io
import os
from datetime import date
from daily_report import (
    load_historical_data, fetch_recent_results, compute_all_states,
    fetch_todays_schedule, get_team_state, build_game_row,
    check_scenarios, API_TO_CANONICAL, SCENARIO_DEFS,
    NEEDS_OPP_STREAK, NEEDS_OPP_ROAD_WP, title_case, fmt_line,
    pl_line_for_trigger, numeric_line, save_odds_to_cache,
)
from master_results_manager import parse_results_upload
from report_builder import build_report_bytes

# Narrow the sidebar
st.markdown("""
<style>
    [data-testid="stSidebar"] {
        min-width: 220px !important;
        max-width: 220px !important;
    }
</style>
""", unsafe_allow_html=True)

@st.cache_data(ttl=600, show_spinner="Loading historical data...")
def load_enriched_data(report_date_str):
    df = load_historical_data()
    df, warnings = fetch_recent_results(df, date.fromisoformat(report_date_str))
    return compute_all_states(df), warnings

@st.cache_data(ttl=600, show_spinner="Fetching today's schedule...")
def get_schedule(report_date_str):
    return fetch_todays_schedule(date.fromisoformat(report_date_str))


def _norm_team_dedup(team):
    """Consistent team label for dedup keys (matches Excel title_case storage)."""
    return title_case(str(team).strip().upper())


def _norm_payout_dedup(val):
    """Consistent payout line string for dedup keys."""
    n = numeric_line(val)
    return str(n) if n is not None else ''


def _master_dedup_key(row, date_str=None):
    """Unique key for a results row; includes opponent + payout line (doubleheader-safe)."""
    if date_str is None:
        try:
            date_str = pd.to_datetime(row.get('Date', '')).strftime('%Y-%m-%d')
        except Exception:
            date_str = str(row.get('Date', '')).strip()[:10]
    opp = str(row.get('_opponent', row.get('Opponent', ''))).strip().upper()
    payout = _norm_payout_dedup(row.get('_line', row.get('PayoutLine', '')))
    return (date_str, _norm_team_dedup(row.get('Team', '')), str(row.get('Scenario', '')), opp, payout)


def _dedup_key_variants(date_str, team, scenario, opponent, payout, *, include_legacy=False):
    """Key forms that count as the same row. Legacy empty-payout keys only for old uploads."""
    team_n = _norm_team_dedup(team)
    scen = str(scenario)
    opp = str(opponent or '').strip().upper()
    pay = _norm_payout_dedup(payout)
    keys = set()
    if pay:
        keys.add((date_str, team_n, scen, opp, pay))
        keys.add((date_str, team_n, scen, '', pay))
    else:
        keys.add((date_str, team_n, scen, opp, ''))
        keys.add((date_str, team_n, scen, '', ''))
    if include_legacy and pay:
        keys.add((date_str, team_n, scen, opp, ''))
        keys.add((date_str, team_n, scen, '', ''))
    return keys


def build_master_file(existing_df, all_triggers, report_date):
    """Build Master_Results.xlsx in memory from prior results + today's triggers."""
    out = io.BytesIO()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Master Results'

    NAVY_FILL  = PatternFill('solid', fgColor='1B2A4A')
    GREEN_FILL = PatternFill('solid', fgColor='375623')
    INPUT_FILL = PatternFill('solid', fgColor='EAF4E8')
    thin = lambda c: Side(style='thin', color=c)
    std_border = Border(left=thin('C6EFCE'), right=thin('C6EFCE'),
                        top=thin('C6EFCE'),  bottom=thin('C6EFCE'))
    input_border = Border(left=thin('375623'), right=thin('375623'),
                          top=thin('375623'),  bottom=thin('375623'))

    for col, width in zip('ABCDEFGHI', [12, 26, 10, 10, 14, 40, 14, 14, 16]):
        ws.column_dimensions[col].width = width

    ws.merge_cells('A1:I1')
    c = ws['A1']
    c.value = 'MASTER RESULTS TRACKER  —  Season Cumulative'
    c.font = Font(bold=True, size=14, color='FFFFFF', name='Calibri')
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:I2')
    c = ws['A2']
    c.value = 'Enter W or L in column H after each game. Net P/L calculates automatically.'
    c.font = Font(italic=True, size=10, color='D0E4F5', name='Calibri')
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16

    headers = ['Date', 'Team', 'H/A', 'Odds', 'Play', 'Scenario', 'Type', 'Result (W/L)', 'Net P/L ($100)', 'PayoutLine', 'Opponent']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col)
        c.value = h
        c.font = Font(bold=True, color='FFFFFF', name='Calibri')
        c.fill = GREEN_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(left=thin('000000'), right=thin('000000'),
                          top=thin('000000'), bottom=thin('000000'))
    ws.row_dimensions[3].height = 24
    ws.freeze_panes = 'A4'
    # Hide helper columns from casual view
    ws.column_dimensions['J'].width = 0.1
    ws.column_dimensions['J'].hidden = True
    ws.column_dimensions['K'].width = 0.1
    ws.column_dimensions['K'].hidden = True

    existing_rows = []
    existing_keys = set()
    today_str = report_date.strftime('%Y-%m-%d')

    if existing_df is not None and not existing_df.empty:
        for _, row in existing_df.iterrows():
            try:
                d = pd.to_datetime(row.get('Date', '')).strftime('%Y-%m-%d')
            except Exception:
                d = str(row.get('Date', '')).strip()[:10]
            if d == today_str and all_triggers:
                continue
            existing_rows.append(row)
            existing_keys.update(_dedup_key_variants(
                d, row.get('Team', ''), row.get('Scenario', ''),
                row.get('_opponent', row.get('Opponent', '')),
                row.get('_line', row.get('PayoutLine', '')),
                include_legacy=not _norm_payout_dedup(row.get('_line', row.get('PayoutLine', ''))),
            ))

    for t in all_triggers:
        scen_str = f"#{t['scenario_id']} {t['scenario']}"
        team_str = title_case(t['team'])
        opp = t.get('opponent', '')
        pl_line = numeric_line(pl_line_for_trigger(t))
        if _dedup_key_variants(today_str, team_str, scen_str, opp, pl_line) & existing_keys:
            continue
        existing_rows.append({
            'Date': today_str,
            'Team': team_str,
            'H/A':  t['home_away'].upper(),
            'Odds': fmt_line(t['line']),
            'Play': t['play'],
            'Scenario': scen_str,
            'Type':   t['verdict'],
            'Result': '',
            'Net P/L': None,
            'PayoutLine': pl_line,
            '_line': pl_line,
            '_opponent': opp,
        })
        existing_keys.update(_dedup_key_variants(today_str, team_str, scen_str, opp, pl_line))

    for i, row in enumerate(existing_rows):
        r = i + 4
        line_val = row.get('_line', None)
        result   = str(row.get('Result', '')).strip().upper()

        ws.cell(r, 1).value = str(row.get('Date', ''))
        ws.cell(r, 2).value = str(row.get('Team', ''))
        ws.cell(r, 3).value = str(row.get('H/A', ''))
        ws.cell(r, 4).value = str(row.get('Odds', ''))
        ws.cell(r, 5).value = str(row.get('Play', ''))
        ws.cell(r, 6).value = str(row.get('Scenario', ''))
        ws.cell(r, 7).value = str(row.get('Type', ''))

        rc = ws.cell(r, 8)
        rc.value = result if result in ('W', 'L') else ''
        rc.fill = INPUT_FILL
        rc.font = Font(bold=True, name='Calibri')
        rc.alignment = Alignment(horizontal='center', vertical='center')
        rc.border = input_border

        # Write P/L as an Excel formula so entering W/L directly in the file auto-calculates
        nc = ws.cell(r, 9)
        # Prefer stored _line (payout line), fall back to PayoutLine col, then Odds
        ln = numeric_line(line_val)
        if ln is None:
            ln = numeric_line(row.get('PayoutLine'))
        if ln is None:
            ln = numeric_line(str(row.get('Odds', '')).replace('+', ''))
        if ln is not None:
            if ln > 0:
                nc.value = f'=IF(H{r}="W",{ln},IF(H{r}="L",-100,""))'
            else:
                nc.value = f'=IF(H{r}="W",ROUND(100/ABS({ln})*100,2),IF(H{r}="L",-100,""))'
        else:
            uploaded_pl = row.get('Net P/L')
            nc.value = uploaded_pl if pd.notna(uploaded_pl) and uploaded_pl != '' else ''
        nc.alignment = Alignment(horizontal='center', vertical='center')
        nc.border = std_border

        # Column J: hidden payout line — persists through upload/download cycle
        ws.cell(r, 10).value = ln
        ws.cell(r, 11).value = str(row.get('_opponent', row.get('Opponent', '')) or '').strip().upper()

        for col in range(1, 8):
            c = ws.cell(r, col)
            c.alignment = Alignment(horizontal='center' if col != 2 else 'left',
                                    vertical='center')
            c.border = std_border

    if existing_rows:
        tr = len(existing_rows) + 4
        ws.merge_cells(f'A{tr}:H{tr}')
        tc = ws[f'A{tr}']
        tc.value = 'TOTAL NET P/L'
        tc.font = Font(bold=True, color='FFFFFF', name='Calibri')
        tc.fill = GREEN_FILL
        tc.alignment = Alignment(horizontal='center', vertical='center')
        tc2 = ws.cell(tr, 9)
        tc2.value = f'=SUMIF(H4:H{tr-1},"W",I4:I{tr-1})+SUMIF(H4:H{tr-1},"L",I4:I{tr-1})'
        tc2.font = Font(bold=True, color='FFFFFF', name='Calibri')
        tc2.fill = GREEN_FILL
        tc2.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ── MAIN UI ───────────────────────────────────────────────────────
st.title("⚾ MLB Daily Betting Report")
st.markdown("---")

col1, col2 = st.columns([2,4])
with col1:
    report_date = st.date_input("Report Date", value=date.today())
report_date_str = report_date.isoformat()

with st.spinner("Loading team data and schedule..."):
    try:
        enriched, api_warnings = load_enriched_data(report_date_str)
        games    = get_schedule(report_date_str)
    except Exception as e:
        st.error(f"Error loading data: {e}")
        st.stop()

for msg in api_warnings:
    st.warning(f"⚠️ Recent results may be incomplete: {msg}")

no_games = not games
if no_games:
    st.warning("No games scheduled for this date.")
else:
    st.success(f"✓ {len(games)} games scheduled for {report_date.strftime('%A, %B %d, %Y')}")
st.markdown("---")

# ── CUMULATIVE TRACKING — upload always visible (even on off-days) ───────
st.subheader("📊 Cumulative Season Tracking")
st.caption(
    "Upload a previous day's saved daily report (with W/L entered on Results Tracker) "
    "or your cumulative Master_Results.xlsx. Then generate today's report."
)

uploaded = st.file_uploader(
    "📤 Upload previous results (Master_Results.xlsx or saved daily report)",
    type=['xlsx'], key='master_upload'
)

c_clear, c_refresh = st.columns(2)
with c_clear:
    if st.button("🗑️ Clear cumulative data", use_container_width=True):
        st.session_state['master_df'] = None
        st.rerun()
with c_refresh:
    if st.button("🔄 Refresh game data cache", use_container_width=True):
        load_enriched_data.clear()
        get_schedule.clear()
        build_heatmap_data.clear()
        st.rerun()

existing_master_df = None
upload_notes = []
if uploaded is not None:
    try:
        existing_master_df, source, sheet, upload_notes = parse_results_upload(uploaded.read())
        st.session_state['master_df'] = existing_master_df
        st.success(
            f"✓ Loaded {len(existing_master_df)} results from {source} "
            f"(sheet: '{sheet}')"
        )
        for note in upload_notes:
            st.warning(note)
    except Exception as e:
        st.warning(f"Could not read uploaded file: {e}. Previous data preserved.")
        # Do NOT clear session — preserve any previously loaded good data
elif 'master_df' not in st.session_state:
    st.session_state['master_df'] = None

# Always read from session state — survives button-click reruns
existing_master_df = st.session_state['master_df']

if existing_master_df is not None and uploaded is None:
    # File was uploaded in a previous interaction this session — show reminder
    st.info(f"📂 Using uploaded Master file ({len(existing_master_df)} rows). Upload a new file to replace it.")

with st.expander("ℹ️ How cumulative tracking works", expanded=False):
    st.markdown("""
    **Simple daily workflow:**
    1. **Upload** either:
       - Your saved **daily report** from yesterday (with W/L entered on Results Tracker), **or**
       - Your cumulative **Master_Results.xlsx** (recommended after day 2+)
    2. **Generate** today's report — new triggers are added automatically, previous results preserved
    3. **Download** the updated `Master_Results.xlsx` at the bottom → enter **W** or **L** for today → save it
    4. **Next day** — upload that Master file (or yesterday's saved daily report) before generating

    Tip: After day 1, always use the **Master_Results.xlsx** download — it accumulates the full season.
    """)

st.markdown("---")

# ── MONEYLINES + GENERATE (only when games are scheduled) ─────────
if no_games:
    st.info("No games scheduled for this date. You can still upload master results above.")
else:
    st.subheader("📋 Enter Today's Moneylines")
    st.caption("Enter the moneyline for each team (e.g. 130 for +130, -150 for -150). Leave blank if unavailable.")

    odds_data = [{'Away Team': g['away_team'], 'Home Team': g['home_team'], 'Away Line': None, 'Home Line': None} for g in games]
    odds_df = pd.DataFrame(odds_data)
    edited = st.data_editor(
        odds_df,
        key=f"odds_{report_date_str}",
        column_config={
            'Away Team': st.column_config.TextColumn('Away Team', disabled=True, width=220),
            'Home Team': st.column_config.TextColumn('Home Team', disabled=True, width=220),
            'Away Line': st.column_config.NumberColumn('Away Line', help='e.g. 130 or -150', width=120),
            'Home Line': st.column_config.NumberColumn('Home Line', help='e.g. -130 or 150', width=120),
        },
        hide_index=True, width='stretch',
        height=(len(odds_df) + 1) * 35 + 3,
    )

    st.markdown("---")

    if st.button("⚾ Generate Daily Report", type="primary", use_container_width=True):
        odds = {}
        odds_skipped = []
        for _, row in edited.iterrows():
            away_team = str(row['Away Team']).upper()
            home_team = str(row['Home Team']).upper()
            if row['Away Line'] is not None and str(row['Away Line']) not in ['', 'nan']:
                try:
                    odds[away_team] = int(float(row['Away Line']))
                except (TypeError, ValueError):
                    odds_skipped.append(f"{away_team}: {row['Away Line']!r}")
            if row['Home Line'] is not None and str(row['Home Line']) not in ['', 'nan']:
                try:
                    odds[home_team] = int(float(row['Home Line']))
                except (TypeError, ValueError):
                    odds_skipped.append(f"{home_team}: {row['Home Line']!r}")

        if odds_skipped:
            st.warning(
                "Could not parse moneyline(s): "
                + ", ".join(odds_skipped)
                + ". Fix or clear those cells and try again."
            )

        if not odds:
            st.warning("Please enter at least some moneylines before generating the report.")
            st.stop()

        save_odds_to_cache(report_date, odds)

        with st.spinner("Running scenario analysis..."):
            opp_streaks = {}; opp_road_wpct = {}
            for team in API_TO_CANONICAL.values():
                tdf = enriched[(enriched['team']==team) & (enriched['date'].dt.date < report_date)]
                if not tdf.empty:
                    last = tdf.iloc[-1]
                    sb = last['streak_before']; res = last['result']
                    opp_streaks[team] = (sb+1 if sb>=0 else 1) if res=='W' else (sb-1 if sb<=0 else -1)
                    road = tdf[tdf['home_away']=='away']
                    if not road.empty:
                        rw=(road['result']=='W').sum(); rl=(road['result']=='L').sum()
                        opp_road_wpct[team] = rw/(rw+rl) if (rw+rl)>0 else None

            all_triggers = []
            missing_history = []
            for g in games:
                away, home = g['away_team'], g['home_team']
                away_state = get_team_state(enriched, away, report_date)
                home_state = get_team_state(enriched, home, report_date)
                if away_state is None:
                    missing_history.append(title_case(away))
                if home_state is None:
                    missing_history.append(title_case(home))
                away_row = build_game_row(away_state, 'away', home, odds.get(away))
                home_row = build_game_row(home_state, 'home', away, odds.get(home))
                all_triggers.extend(check_scenarios([away_row, home_row], opp_streaks, opp_road_wpct, odds))

            if missing_history:
                st.warning(
                    "No historical data for: "
                    + ", ".join(sorted(set(missing_history)))
                    + " — scenario checks skipped for those teams."
                )

            master_bytes = build_master_file(existing_master_df, all_triggers, report_date)
            merged_df = None
            try:
                merged_df, _, _, _ = parse_results_upload(master_bytes)
                st.session_state['master_df'] = merged_df
            except Exception as e:
                st.warning(
                    f"Master file was built but session state could not be refreshed "
                    f"({e}). Re-upload the downloaded Master_Results.xlsx before generating again."
                )

            excel_bytes, n_total, n_fade, n_inc = build_report_bytes(
                games, all_triggers, report_date, odds, cumulative_df=merged_df,
            )

        fade_no_opp = [t for t in all_triggers if t['verdict'] == 'CLEAR FADE' and t.get('opp_line') is None]
        if fade_no_opp:
            st.warning(
                f"{len(fade_no_opp)} CLEAR FADE trigger(s) are missing opponent odds — "
                "Net P/L is left blank until the opponent moneyline is entered."
            )

        st.markdown("---")
        st.subheader("📊 Results")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total Games", len(games))
        m2.metric("🟢 Clear Bet", sum(1 for t in all_triggers if t['verdict']=='CLEAR BET'))
        m3.metric("🔴 Clear Fade", sum(1 for t in all_triggers if t['verdict']=='CLEAR FADE'))
        m4.metric("🟡 Inconsistent", sum(1 for t in all_triggers if t['verdict']=='INCONSISTENT'))

        for section_label, section_verdict in [
            ("🟢 Clear Bet Plays", "CLEAR BET"),
            ("🔴 Clear Fade Plays", "CLEAR FADE"),
            ("🟡 Inconsistent Plays", "INCONSISTENT"),
        ]:
            section_triggers = [t for t in all_triggers if t['verdict']==section_verdict]
            if not section_triggers: continue
            st.markdown(f"### {section_label}")
            game_rows = []
            seen = set()
            for g in games:
                away, home = g['away_team'], g['home_team']
                key = f"{away}@{home}"
                if key in seen: continue
                at = [t for t in section_triggers if t['team']==away and t['opponent']==home]
                ht = [t for t in section_triggers if t['team']==home and t['opponent']==away]
                if not at and not ht: continue
                seen.add(key)
                matchup = f"{title_case(away)} @ {title_case(home)}"
                game_rows.append({'GAME': matchup, 'H/A': 'AWAY', 'Team': title_case(away),
                    'Odds': fmt_line(odds.get(away)),
                    'Play': at[0]['play'] if at else '',
                    'Scenario': f"#{at[0]['scenario_id']} {at[0]['scenario']}" if at else ''})
                game_rows.append({'GAME': '', 'H/A': 'HOME', 'Team': title_case(home),
                    'Odds': fmt_line(odds.get(home)),
                    'Play': ht[0]['play'] if ht else '',
                    'Scenario': f"#{ht[0]['scenario_id']} {ht[0]['scenario']}" if ht else ''})
            if game_rows:
                n_rows = len(game_rows)
                st.dataframe(pd.DataFrame(game_rows),
                    width='stretch', hide_index=True,
                    height=(n_rows + 1) * 35 + 3,
                    column_config={
                        'GAME': st.column_config.TextColumn('GAME', width=300),
                        'H/A':  st.column_config.TextColumn('H/A',  width=60),
                        'Team': st.column_config.TextColumn('Team', width=180),
                        'Odds': st.column_config.TextColumn('Odds', width=70),
                        'Play': st.column_config.TextColumn('Play', width=200),
                        'Scenario': st.column_config.TextColumn('Scenario', width=320),
                    })

        st.markdown("---")
        fname = f'MLB_Daily_Report_{report_date.strftime("%Y-%m-%d")}.xlsx'
        st.download_button(label="📥 Download Excel Report", data=excel_bytes, file_name=fname,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True, type="primary")

        st.markdown("---")
        st.subheader("📊 Download Updated Master Results")
        st.download_button(
            label="📥 Download Master_Results.xlsx  (open → enter W/L → re-upload next time)",
            data=master_bytes,
            file_name='Master_Results.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True,
        )


# ── SCENARIO PERFORMANCE HEATMAP ─────────────────────────────────
st.markdown("---")
st.subheader("🔥 Scenario Performance Heatmap  (2023–2026 Historical Backtest)")
st.caption("Win % color-coded: 🟢 green = strong, 🟡 yellow = mixed, 🔴 red = weak. INCONSISTENT scenarios show N/A (no defined bet direction). Click column headers to sort.")

@st.cache_data(ttl=3600, show_spinner="Building heatmap from historical data...")
def build_heatmap_data():
    """
    Run all 36 scenarios against the full historical dataset using fast
    vectorized pandas — no row-by-row iteration.
    Returns a summary DataFrame with W, L, Win% per scenario.
    """
    from daily_report import (
        load_historical_data, compute_all_states, SCENARIO_DEFS,
        NEEDS_OPP_STREAK, NEEDS_OPP_ROAD_WP,
        build_streak_lookup, build_opp_road_wpct_lookup, backtest_wl_counts,
    )

    df = load_historical_data()
    df = compute_all_states(df)
    df = df[df['result'].isin(['W', 'L'])].copy()

    streak_lookup = build_streak_lookup(df)
    opp_road_lookup = build_opp_road_wpct_lookup(df)

    rows = []
    for sid, sname, verdict, func in SCENARIO_DEFS:
        try:
            if sid in NEEDS_OPP_STREAK:
                mask = df.apply(
                    lambda r: bool(func(
                        r.to_dict(),
                        {r['opponent']: streak_lookup.get((r['date'], r['opponent']), 0)},
                    )),
                    axis=1,
                )
            elif sid in NEEDS_OPP_ROAD_WP:
                mask = df.apply(
                    lambda r: bool(func(
                        r.to_dict(),
                        {r['opponent']: opp_road_lookup.get((r['date'], r['opponent']))},
                    )),
                    axis=1,
                )
            else:
                mask = df.apply(lambda r: bool(func(r.to_dict())), axis=1)
        except Exception:
            mask = pd.Series(False, index=df.index)

        subset = df[mask]
        total  = len(subset)
        wins, losses = backtest_wl_counts(subset, verdict)
        if wins is None:
            win_pct = None
            w_disp = l_disp = None
        else:
            win_pct = round(wins / total * 100, 1) if total > 0 else None
            w_disp, l_disp = wins, losses

        rows.append({
            '#':        sid,
            'Scenario': sname,
            'Type':     verdict,
            'Games':    total,
            'W':        w_disp,
            'L':        l_disp,
            'Win%':     win_pct,
        })

    return pd.DataFrame(rows)

with st.expander("Click to view heatmap", expanded=False):
    with st.spinner("Running scenarios against historical data… (first load ~30 sec, then cached)"):
        hmap = build_heatmap_data()

    if not hmap.empty:
        c1, c2 = st.columns([2, 2])
        with c1:
            sort_opt = st.selectbox(
                "Sort by",
                ["Win% High→Low", "Win% Low→High", "Scenario #", "Games High→Low"],
                key='hm_sort'
            )
        with c2:
            type_filter = st.multiselect(
                "Type",
                ['CLEAR BET', 'CLEAR FADE', 'INCONSISTENT'],
                default=['CLEAR BET', 'CLEAR FADE', 'INCONSISTENT'],
                key='hm_type'
            )

        hmap = hmap[hmap['Type'].isin(type_filter)].copy()

        if sort_opt == "Win% High→Low":
            hmap = hmap.sort_values('Win%', ascending=False, na_position='last')
        elif sort_opt == "Win% Low→High":
            hmap = hmap.sort_values('Win%', ascending=True, na_position='last')
        elif sort_opt == "Games High→Low":
            hmap = hmap.sort_values('Games', ascending=False)
        else:
            hmap = hmap.sort_values('#')

        def _row_style(row):
            # Always set explicit dark text so colors are visible in both light and dark mode
            BASE = 'color:#000000;font-weight:bold'
            v = row['Win%']
            vtype = row['Type']
            try:
                v = float(v) if v is not None else None
            except (TypeError, ValueError):
                v = None
            if v is None:
                wc = f'background-color:#D0D0D0;{BASE}'
            elif vtype == 'CLEAR FADE':
                wc = (f'background-color:#C6EFCE;{BASE}' if v <= 42
                      else f'background-color:#FFEB9C;{BASE}' if v <= 50
                      else f'background-color:#FFC7CE;{BASE}')
            else:
                wc = (f'background-color:#C6EFCE;{BASE}' if v >= 58
                      else f'background-color:#FFEB9C;{BASE}' if v >= 50
                      else f'background-color:#FFC7CE;{BASE}')
            tc = {
                'CLEAR BET':    f'background-color:#C6EFCE;{BASE}',
                'CLEAR FADE':   f'background-color:#FFC7CE;{BASE}',
                'INCONSISTENT': f'background-color:#FFEB9C;{BASE}',
            }.get(vtype, '')
            plain = f'color:#000000'
            return [plain, plain, tc, plain,
                    f'background-color:#E2EFDA;{BASE}',
                    f'background-color:#FFE7E7;{BASE}',
                    wc]

        # Apply styles on numeric data first, then format Win% as string for display
        disp = hmap.copy()
        styled = disp.style.apply(_row_style, axis=1).format({'Win%': lambda x: f"{x:.1f}%" if x is not None else "—"})

        st.dataframe(
            styled,
            use_container_width=True,
            hide_index=True,
            height=(len(hmap) + 1) * 35 + 3,
            column_config={
                '#':        st.column_config.TextColumn('#',        width=40),
                'Scenario': st.column_config.TextColumn('Scenario', width=260),
                'Type':     st.column_config.TextColumn('Type',     width=120),
                'Games':    st.column_config.NumberColumn('Games',  width=60),
                'W':        st.column_config.NumberColumn('W',      width=45),
                'L':        st.column_config.NumberColumn('L',      width=45),
                'Win%':     st.column_config.TextColumn('Win %',    width=65),
            }
        )

        has = hmap[hmap['Games'] > 0].copy()
        if not has.empty and has['Win%'].notna().any():
            best  = has.loc[has['Win%'].idxmax()]
            worst = has.loc[has['Win%'].idxmin()]
            b1, b2, b3 = st.columns(3)
            b1.success(f"🏆 Best: #{best['#']} — {best['Win%']:.1f}%")
            b2.error(f"⚠️ Lowest: #{worst['#']} — {worst['Win%']:.1f}%")
            b3.info(f"📊 {len(has)} scenarios · {int(hmap['Games'].sum()):,} total games")
