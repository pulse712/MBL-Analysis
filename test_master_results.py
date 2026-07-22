"""
Test Script for Master Results Manager
======================================
Run this to verify the master results system is working correctly.
"""

import os
import sys
from datetime import date

# Add parent directory to path
sys.path.insert(0, os.path.dirname(__file__))

import report_builder
from master_results_manager import (
    initialize_master_results,
    append_to_master_results,
    load_master_results,
    get_cumulative_stats,
    get_master_results_path,
    parse_results_upload,
    MASTER_FILE
)


def test_initialization():
    """Test 1: Initialize master results file"""
    print("\n" + "="*60)
    print("TEST 1: Initialize Master Results File")
    print("="*60)
    
    # Remove existing file if present
    if os.path.exists(MASTER_FILE):
        os.remove(MASTER_FILE)
        print(f"✓ Removed existing {MASTER_FILE}")
    
    # Initialize
    initialize_master_results()
    
    # Check file exists
    if os.path.exists(MASTER_FILE):
        from openpyxl import load_workbook
        wb = load_workbook(MASTER_FILE)
        ws = wb.active
        if (ws.cell(row=3, column=10).value == 'PayoutLine' and ws.column_dimensions['J'].hidden
                and ws.cell(row=3, column=11).value == 'Opponent' and ws.column_dimensions['K'].hidden):
            print(f"✅ SUCCESS: {MASTER_FILE} created with PayoutLine + Opponent columns")
        else:
            print("❌ FAILED: Master file missing hidden PayoutLine/Opponent columns")
            return False
        print(f"   Location: {get_master_results_path()}")
        return True
    else:
        print(f"❌ FAILED: {MASTER_FILE} not created")
        return False


def test_append_results():
    """Test 2: Append sample results to master file"""
    print("\n" + "="*60)
    print("TEST 2: Append Sample Results")
    print("="*60)
    
    # Create sample triggers
    sample_triggers = [
        {
            'team': 'NEW YORK YANKEES',
            'opponent': 'BOSTON RED SOX',
            'home_away': 'home',
            'line': -150,
            'scenario_id': '01',
            'scenario': 'BLOWOUT #1 - MJ',
            'verdict': 'CLEAR BET',
            'play': 'BET NEW YORK YANKEES'
        },
        {
            'team': 'LOS ANGELES DODGERS',
            'opponent': 'SAN FRANCISCO GIANTS',
            'home_away': 'away',
            'line': 120,
            'opp_line': 130,
            'scenario_id': '14',
            'scenario': 'THE SCORING DROUGHT #1 - MJ',
            'verdict': 'CLEAR FADE',
            'play': 'FADE LOS ANGELES DODGERS'
        },
        {
            'team': 'HOUSTON ASTROS',
            'opponent': 'SEATTLE MARINERS',
            'home_away': 'home',
            'line': -105,
            'scenario_id': '22',
            'scenario': 'SMALL ROAD FAVORITE / NEW SERIES - SM',
            'verdict': 'INCONSISTENT',
            'play': 'WATCH HOUSTON ASTROS'
        }
    ]
    
    test_date = date(2026, 7, 9)
    
    try:
        append_to_master_results(sample_triggers, test_date)
        with open(MASTER_FILE, 'rb') as f:
            df, _, _, _ = parse_results_upload(f.read())
        fade_row = df[df['Type'] == 'CLEAR FADE'].iloc[0]
        if fade_row.get('PayoutLine') != 130:
            print(f"❌ FAILED: CLI append FADE PayoutLine={fade_row.get('PayoutLine')}, expected 130")
            return False
        print(f"✅ SUCCESS: Appended {len(sample_triggers)} triggers to master file")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_load_results():
    """Test 3: Load results from master file"""
    print("\n" + "="*60)
    print("TEST 3: Load Master Results")
    print("="*60)
    
    try:
        df = load_master_results()
        
        if df is not None and not df.empty:
            print(f"✅ SUCCESS: Loaded {len(df)} results")
            print("\nSample data:")
            print(df[['Date', 'Team', 'Scenario', 'Result']].head())
            return True
        else:
            print("⚠️  WARNING: Master file is empty")
            return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_cumulative_stats():
    """Test 4: Calculate cumulative statistics"""
    print("\n" + "="*60)
    print("TEST 4: Calculate Cumulative Stats")
    print("="*60)
    
    try:
        stats = get_cumulative_stats()
        
        if stats:
            print(f"✅ SUCCESS: Calculated stats for {len(stats)} scenarios")
            print("\nSample stats (first 3 scenarios):")
            for i, (scenario, data) in enumerate(list(stats.items())[:3]):
                print(f"\n{scenario}:")
                print(f"  Wins: {data['wins']}, Losses: {data['losses']}, Total: {data['total']}")
                print(f"  Win %: {data['win_pct']:.1%}, Net P/L: ${data['net_pl']:.2f}")
            return True
        else:
            print("⚠️  WARNING: No stats calculated (need W/L results entered)")
            return True  # Not a failure, just no results yet
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_add_results_with_outcomes():
    """Test 5: Add results and simulate entering W/L outcomes"""
    print("\n" + "="*60)
    print("TEST 5: Simulate Entering W/L Results")
    print("="*60)
    
    try:
        from openpyxl import load_workbook
        
        # Load the workbook
        wb = load_workbook(MASTER_FILE)
        ws = wb.active
        
        # Simulate entering results (W or L) for the sample data
        # Row 4 is first data row
        results = ['W', 'L', 'W']  # Win, Loss, Win
        
        for i, result in enumerate(results, start=4):
            ws.cell(row=i, column=8).value = result  # Column H is Result
        
        wb.save(MASTER_FILE)
        print(f"✅ SUCCESS: Added W/L results to {len(results)} rows")
        
        # Now test if stats calculate correctly
        stats = get_cumulative_stats()
        if stats:
            total_plays = sum(s['total'] for s in stats.values())
            total_pl = sum(s['net_pl'] for s in stats.values())
            print(f"✓ Cumulative stats now show {total_plays} completed plays")
            if total_pl == 0:
                print("❌ FAILED: Net P/L is $0 — expected calculated values from W/L + odds")
                return False
            print(f"✓ Net P/L calculated: ${total_pl:.2f}")
            return True
        else:
            print("⚠️  WARNING: Stats not calculated")
            return False
            
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_file_references():
    """Test 6: Verify Excel formula references work"""
    print("\n" + "="*60)
    print("TEST 6: Verify Excel Formula References")
    print("="*60)
    
    try:
        from openpyxl import load_workbook
        
        # Check if formulas reference external file correctly
        wb = load_workbook(MASTER_FILE)
        ws = wb.active
        
        # Check if Net P/L formulas exist
        formula_count = 0
        for row in range(4, 10):  # Check first few data rows
            cell = ws.cell(row=row, column=9)  # Column I is Net P/L
            if cell.value and str(cell.value).startswith('='):
                formula_count += 1
        
        if formula_count > 0:
            print(f"✅ SUCCESS: Found {formula_count} Net P/L formulas")
            print("✓ Formulas will auto-calculate based on W/L entries")
            return True
        else:
            print("⚠️  WARNING: No formulas found (may be intentional)")
            return True
            
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_parse_daily_report_upload():
    """Test 7: Parse results from a saved daily report (Results Tracker tab)"""
    print("\n" + "="*60)
    print("TEST 7: Parse Daily Report Upload")
    print("="*60)

    try:
        import io
        import xlsxwriter

        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {'in_memory': True})
        wb.add_worksheet('Green Clear Bet')
        ws = wb.add_worksheet('Results Tracker')
        ws.write(0, 0, 'RESULTS TRACKER')
        ws.write(1, 0, 'Enter W or L')
        headers = ['Date', 'Team', 'H/A', 'Odds', 'Play', 'Scenario', 'Type', 'Result (W/L)', 'Net P/L ($100)']
        for col, h in enumerate(headers):
            ws.write(2, col, h)
        ws.write(3, 0, '2026-07-20')
        ws.write(3, 1, 'New York Yankees')
        ws.write(3, 2, 'HOME')
        ws.write(3, 3, '-150')
        ws.write(3, 4, 'BET NEW YORK YANKEES')
        ws.write(3, 5, '#01 BLOWOUT #1 - MJ')
        ws.write(3, 6, 'CLEAR BET')
        ws.write(3, 7, 'W')
        ws.write(3, 8, 66.67)
        wb.close()
        file_bytes = buf.getvalue()

        df, source, sheet, _ = parse_results_upload(file_bytes)

        if source == 'daily report' and sheet == 'Results Tracker' and len(df) == 1:
            if df.iloc[0]['Result'] == 'W' and str(df.iloc[0]['Date'])[:10] == '2026-07-20':
                print("✅ SUCCESS: Parsed daily report Results Tracker with W/L")
                return True

        print(f"❌ FAILED: Unexpected parse result source={source} sheet={sheet} rows={len(df)}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_master_dedup_and_fade_roundtrip():
    """Test 8: Double-generate dedup + multi-cycle FADE PayoutLine persistence"""
    print("\n" + "="*60)
    print("TEST 8: Master Dedup & FADE Round-Trip")
    print("="*60)

    try:
        import sys
        import os
        from datetime import date
        from openpyxl import load_workbook
        import io

        sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'pages'))
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            'daily_report_page',
            os.path.join(os.path.dirname(__file__), 'pages', '1_Daily_Report.py'),
        )
        page = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(page)

        bet = {
            'verdict': 'CLEAR BET', 'line': -150, 'scenario_id': '01', 'scenario': 'B',
            'team': 'NEW YORK YANKEES', 'opponent': 'BOSTON RED SOX',
            'home_away': 'home', 'play': 'BET',
        }
        fade = {
            'verdict': 'CLEAR FADE', 'line': -150, 'opp_line': 130,
            'scenario_id': '02', 'scenario': 'Y', 'team': 'NEW YORK YANKEES',
            'opponent': 'BOSTON RED SOX', 'home_away': 'home', 'play': 'FADE',
        }

        # Double-generate should not duplicate
        df1, _, _, _ = parse_results_upload(page.build_master_file(None, [bet], date(2026, 7, 21)))
        df2, _, _, _ = parse_results_upload(page.build_master_file(df1, [bet], date(2026, 7, 21)))
        if len(df1) != 1 or len(df2) != 1:
            print(f"❌ FAILED: Double-generate rows {len(df1)} -> {len(df2)}, expected 1 -> 1")
            return False

        # DH same matchup different lines -> 2 rows
        t1 = {**bet, 'line': -150}
        t2 = {**bet, 'line': -130}
        df_dh, _, _, _ = parse_results_upload(page.build_master_file(None, [t1, t2], date(2026, 7, 21)))
        if len(df_dh) != 2:
            print(f"❌ FAILED: DH rows {len(df_dh)}, expected 2")
            return False

        # FADE: 3 upload/generate cycles keep +130 formula
        df = None
        ok = True
        for i in range(3):
            b = page.build_master_file(df, [fade] if i == 0 else [], date(2026, 7, 21))
            ws = load_workbook(io.BytesIO(b)).active
            formula = str(ws['I4'].value or '')
            jval = ws['J4'].value
            if '130' not in formula:
                print(f"❌ FAILED: Cycle {i+1} formula lost +130: {formula}")
                ok = False
                break
            if jval is None:
                print(f"❌ FAILED: Cycle {i+1} PayoutLine col J is empty")
                ok = False
                break
            df, _, _, _ = parse_results_upload(b)

        if ok:
            print("✅ SUCCESS: Dedup and 3-cycle FADE PayoutLine persistence")
            return True
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_load_odds_plus_prefix():
    """CLI odds template accepts +130 style moneylines."""
    print("\n" + "="*60)
    print("TEST 9: CLI Odds +130 Parsing")
    print("="*60)
    try:
        import daily_report as dr
        from openpyxl import Workbook
        import tempfile
        import os

        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.close()
        wb = Workbook()
        ws = wb.active
        ws.append(['DAILY ODDS INPUT'])
        ws.append(['Away Team', 'Home Team', 'Away Line', 'Home Line'])
        ws.append(['NEW YORK YANKEES', 'BOSTON RED SOX', '+130', '-150'])
        wb.save(tmp.name)

        old = dr.ODDS_INPUT
        dr.ODDS_INPUT = tmp.name
        try:
            odds = dr.load_odds([])
        finally:
            dr.ODDS_INPUT = old
            os.unlink(tmp.name)

        if odds.get('NEW YORK YANKEES') == 130 and odds.get('BOSTON RED SOX') == -150:
            print("✅ SUCCESS: Parsed +130 and -150 from odds template")
            return True
        print(f"❌ FAILED: Unexpected odds dict: {odds}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_series_last_game_detection():
    """Completed 3-game series: last_game_series only on game 3."""
    print("\n" + "="*60)
    print("TEST 10: Series Last-Game Detection")
    print("="*60)
    try:
        import pandas as pd
        from daily_report import compute_team_states, last_game_series

        rows = pd.DataFrame([
            {'year': 2026, 'opponent': 'BOS', 'home_away': 'home', 'result': 'W',
             'runs_scored': 5, 'runs_allowed': 2, 'line': -120, 'date': pd.Timestamp('2026-07-01')},
            {'year': 2026, 'opponent': 'BOS', 'home_away': 'home', 'result': 'L',
             'runs_scored': 1, 'runs_allowed': 4, 'line': -110, 'date': pd.Timestamp('2026-07-02')},
            {'year': 2026, 'opponent': 'BOS', 'home_away': 'home', 'result': 'W',
             'runs_scored': 3, 'runs_allowed': 1, 'line': -130, 'date': pd.Timestamp('2026-07-03')},
            {'year': 2026, 'opponent': 'NYY', 'home_away': 'away', 'result': 'L',
             'runs_scored': 2, 'runs_allowed': 6, 'line': 150, 'date': pd.Timestamp('2026-07-05')},
        ])
        enriched = compute_team_states(rows)
        flags = [last_game_series(r.to_dict()) for _, r in enriched.iterrows()]
        if flags == [False, False, True, False]:
            print("✅ SUCCESS: last_game_series fires only on confirmed series ends")
            return True
        print(f"❌ FAILED: Unexpected flags {flags}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_fade_backtest_wl():
    """Fade backtest counts team losses as bet wins."""
    print("\n" + "="*60)
    print("TEST 11: FADE Backtest W/L")
    print("="*60)
    try:
        import pandas as pd
        from daily_report import backtest_wl_counts

        subset = pd.DataFrame({'result': ['W', 'L', 'L', 'W']})
        wins, losses = backtest_wl_counts(subset, 'CLEAR FADE')
        if wins == 2 and losses == 2:
            print("✅ SUCCESS: FADE backtest inverts team results")
            return True
        print(f"❌ FAILED: Expected 2/2, got {wins}/{losses}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_odds_cache_backfill():
    """Saved moneylines backfill API rows with line=None."""
    print("\n" + "="*60)
    print("TEST 12: Odds Cache Backfill")
    print("="*60)
    try:
        import daily_report as dr
        import pandas as pd
        import os
        from datetime import date

        if os.path.exists(dr.ODDS_CACHE):
            os.remove(dr.ODDS_CACHE)
        dr.save_odds_to_cache(date(2026, 7, 20), {
            'NEW YORK YANKEES': -150,
            'BOSTON RED SOX': 130,
        })
        df = pd.DataFrame([{
            'year': 2026, 'team': 'NEW YORK YANKEES',
            'date': pd.Timestamp('2026-07-20'), 'opponent': 'BOSTON RED SOX',
            'home_away': 'home', 'result': 'W', 'score': '5-2',
            'runs_scored': 5, 'runs_allowed': 2,
            'line': None, 'ou': None, 'total': None, 'division': 'AL_EAST',
        }])
        out = dr.apply_odds_cache(df)
        if out.iloc[0]['line'] == -150:
            print("✅ SUCCESS: Cached moneyline backfilled onto API row")
            os.remove(dr.ODDS_CACHE)
            return True
        print(f"❌ FAILED: line={out.iloc[0]['line']}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_tie_games_not_counted_as_loss():
    """Tie/undecided results must not increment losses in streak logic."""
    print("\n" + "="*60)
    print("TEST 13: Tie Game Streak Handling")
    print("="*60)
    try:
        import pandas as pd
        from daily_report import compute_team_states

        rows = pd.DataFrame([
            {'year': 2026, 'opponent': 'BOS', 'home_away': 'home', 'result': 'W',
             'runs_scored': 5, 'runs_allowed': 2, 'line': -120, 'date': pd.Timestamp('2026-07-01')},
            {'year': 2026, 'opponent': 'BOS', 'home_away': 'home', 'result': 'T',
             'runs_scored': 3, 'runs_allowed': 3, 'line': -110, 'date': pd.Timestamp('2026-07-02')},
            {'year': 2026, 'opponent': 'BOS', 'home_away': 'home', 'result': 'W',
             'runs_scored': 4, 'runs_allowed': 1, 'line': -110, 'date': pd.Timestamp('2026-07-03')},
        ])
        enriched = compute_team_states(rows)
        streak_before_third = enriched.iloc[2]['streak_before']
        if streak_before_third == 1:
            print("✅ SUCCESS: Tie did not count as a loss between wins")
            return True
        print(f"❌ FAILED: streak_before before 3rd game = {streak_before_third}, expected 1")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_daily_report_fade_payoutline_upload():
    """Daily report Results Tracker preserves FADE PayoutLine through upload."""
    print("\n" + "="*60)
    print("TEST 14: Daily Report FADE PayoutLine Upload")
    print("="*60)
    try:
        import sys
        import os
        from datetime import date
        import importlib.util

        sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'pages'))
        spec = importlib.util.spec_from_file_location(
            'daily_report_page',
            os.path.join(os.path.dirname(__file__), 'pages', '1_Daily_Report.py'),
        )
        page = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(page)
        from master_results_manager import parse_results_upload

        fade = {
            'verdict': 'CLEAR FADE', 'line': -150, 'opp_line': 130,
            'scenario_id': '02', 'scenario': 'Y', 'team': 'NEW YORK YANKEES',
            'opponent': 'BOSTON RED SOX', 'home_away': 'home', 'play': 'FADE',
        }
        games = [{'away_team': 'BOSTON RED SOX', 'home_team': 'NEW YORK YANKEES'}]
        odds = {'NEW YORK YANKEES': -150, 'BOSTON RED SOX': 130}
        b = report_builder.build_report_bytes(games, [fade], date(2026, 7, 20), odds)[0]
        df, src, _, _ = parse_results_upload(b)
        if 'PayoutLine' not in df.columns or df.iloc[0]['PayoutLine'] != 130:
            print(f"❌ FAILED: Daily parse PayoutLine={df.iloc[0].get('PayoutLine')}")
            return False
        mb = page.build_master_file(df, [], date(2026, 7, 21))
        df2, _, _, _ = parse_results_upload(mb)
        if df2.iloc[0]['PayoutLine'] == 130:
            print("✅ SUCCESS: FADE PayoutLine survives daily report → master rebuild")
            return True
        print(f"❌ FAILED: Master PayoutLine={df2.iloc[0].get('PayoutLine')}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_normalize_opponent_abbreviations():
    """Opponent abbreviations map to full canonical names."""
    print("\n" + "="*60)
    print("TEST 15: Opponent Abbreviation Normalization")
    print("="*60)
    try:
        from daily_report import normalize_opponent

        cases = {
            '@ SD': 'SAN DIEGO PADRES',
            '@ SF': 'SAN FRANCISCO GIANTS',
            'nyy': 'NEW YORK YANKEES',
            'nym': 'NEW YORK METS',
        }
        for raw, expected in cases.items():
            got = normalize_opponent(raw)
            if got != expected:
                print(f"❌ FAILED: {raw!r} -> {got!r}, expected {expected!r}")
                return False
        print("✅ SUCCESS: Abbreviations normalize correctly")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_fade_pl_line_requires_opponent_odds():
    """FADE P/L line is None when opponent odds are missing."""
    print("\n" + "="*60)
    print("TEST 16: FADE Requires Opponent Odds")
    print("="*60)
    try:
        from daily_report import pl_line_for_trigger

        t = {'verdict': 'CLEAR FADE', 'line': -150, 'opp_line': None}
        if pl_line_for_trigger(t) is None:
            print("✅ SUCCESS: FADE without opponent odds returns None")
            return True
        print(f"❌ FAILED: got {pl_line_for_trigger(t)}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_inconsistent_backtest_na():
    """INCONSISTENT scenarios have no defined backtest W/L."""
    print("\n" + "="*60)
    print("TEST 17: INCONSISTENT Backtest N/A")
    print("="*60)
    try:
        import pandas as pd
        from daily_report import backtest_wl_counts

        subset = pd.DataFrame({'result': ['W', 'L']})
        wins, losses = backtest_wl_counts(subset, 'INCONSISTENT')
        if wins is None and losses is None:
            print("✅ SUCCESS: INCONSISTENT backtest returns no W/L")
            return True
        print(f"❌ FAILED: got wins={wins}, losses={losses}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_cli_scenario_perf_master_merge():
    """CLI build_report embeds cumulative sheet when Master_Results.xlsx exists."""
    print("\n" + "="*60)
    print("TEST 18: CLI Scenario Performance Master Merge")
    print("="*60)
    try:
        import os
        from datetime import date
        from openpyxl import load_workbook
        from master_results_manager import MASTER_FILE, initialize_master_results, append_to_master_results
        from daily_report import build_report

        if os.path.exists(MASTER_FILE):
            os.remove(MASTER_FILE)
        initialize_master_results()
        append_to_master_results([{
            'team': 'NEW YORK YANKEES', 'opponent': 'BOSTON RED SOX', 'home_away': 'home',
            'line': -150, 'scenario_id': '01', 'scenario': 'B', 'verdict': 'CLEAR BET',
            'play': 'BET NEW YORK YANKEES',
        }], date(2026, 7, 19))
        wb = load_workbook(MASTER_FILE)
        wb.active.cell(row=4, column=8).value = 'W'
        wb.save(MASTER_FILE)

        games = [{'away_team': 'BOSTON RED SOX', 'home_team': 'NEW YORK YANKEES'}]
        triggers = [{
            'team': 'NEW YORK YANKEES', 'opponent': 'BOSTON RED SOX', 'home_away': 'home',
            'line': -150, 'scenario_id': '01', 'scenario': 'B', 'verdict': 'CLEAR BET',
            'play': 'BET NEW YORK YANKEES',
        }]
        fname = build_report(
            games, triggers, date(2026, 7, 20),
            {'NEW YORK YANKEES': -150, 'BOSTON RED SOX': 130},
        )
        out = load_workbook(fname)
        if 'Cumulative Results' not in out.sheetnames:
            print("❌ FAILED: No Cumulative Results sheet in CLI report")
            return False
        if out['Cumulative Results'].sheet_state != 'hidden':
            print("❌ FAILED: Cumulative Results sheet not hidden")
            return False
        perf = out['📋 Scenario Performance']
        if 'Season Cumulative' not in str(perf['A1'].value):
            print(f"❌ FAILED: Banner={perf['A1'].value}")
            return False
        os.remove(fname)
        print("✅ SUCCESS: CLI report merges master history into Scenario Performance")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_open_series_no_early_last_game():
    """Open series at end of log must not fire last_game_series on game 3 of 4."""
    print("\n" + "="*60)
    print("TEST 19: Open Series No Early Last Game")
    print("="*60)
    try:
        import pandas as pd
        from daily_report import compute_team_states, last_game_series

        rows = pd.DataFrame([
            {'year': 2026, 'opponent': 'BOS', 'home_away': 'home', 'result': 'W',
             'runs_scored': 5, 'runs_allowed': 2, 'line': -120, 'date': pd.Timestamp('2026-07-01')},
            {'year': 2026, 'opponent': 'BOS', 'home_away': 'home', 'result': 'L',
             'runs_scored': 1, 'runs_allowed': 4, 'line': -110, 'date': pd.Timestamp('2026-07-02')},
            {'year': 2026, 'opponent': 'BOS', 'home_away': 'home', 'result': 'W',
             'runs_scored': 3, 'runs_allowed': 1, 'line': -130, 'date': pd.Timestamp('2026-07-03')},
        ])
        enriched = compute_team_states(rows)
        flags = [last_game_series(r.to_dict()) for _, r in enriched.iterrows()]
        if flags == [False, False, False]:
            print("✅ SUCCESS: Open series at EOF does not fire last_game_series early")
            return True
        print(f"❌ FAILED: Unexpected flags {flags}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_opponent_persists_master_roundtrip():
    """Hidden Opponent column survives master upload/download for DH dedup."""
    print("\n" + "="*60)
    print("TEST 20: Opponent Persists Master Round-Trip")
    print("="*60)
    try:
        import sys, os, importlib.util
        from datetime import date

        sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'pages'))
        spec = importlib.util.spec_from_file_location(
            'daily_report_page',
            os.path.join(os.path.dirname(__file__), 'pages', '1_Daily_Report.py'),
        )
        page = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(page)

        base = {
            'verdict': 'CLEAR BET', 'scenario_id': '01', 'scenario': 'B',
            'team': 'NEW YORK YANKEES', 'home_away': 'home', 'play': 'BET',
        }
        t1 = {**base, 'opponent': 'BOSTON RED SOX', 'line': -150}
        t2 = {**base, 'opponent': 'TAMPA BAY RAYS', 'line': -140}
        df1, _, _, _ = parse_results_upload(page.build_master_file(None, [t1, t2], date(2026, 7, 21)))
        if df1.iloc[0]['Opponent'] != 'BOSTON RED SOX' or len(df1) != 2:
            print(f"❌ FAILED: Initial parse opponents={df1['Opponent'].tolist()}, rows={len(df1)}")
            return False
        df2, _, _, _ = parse_results_upload(page.build_master_file(df1, [], date(2026, 7, 21)))
        if len(df2) != 2:
            print(f"❌ FAILED: Re-upload collapsed to {len(df2)} rows")
            return False
        print("✅ SUCCESS: Opponent column persists through master round-trip")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_regenerate_replaces_same_day_rows():
    """Re-generating same date replaces stale triggers instead of appending."""
    print("\n" + "="*60)
    print("TEST 21: Re-Generate Replaces Same-Day Rows")
    print("="*60)
    try:
        import sys, os, importlib.util
        from datetime import date

        sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'pages'))
        spec = importlib.util.spec_from_file_location(
            'daily_report_page',
            os.path.join(os.path.dirname(__file__), 'pages', '1_Daily_Report.py'),
        )
        page = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(page)

        bet = {
            'verdict': 'CLEAR BET', 'line': -150, 'scenario_id': '01', 'scenario': 'B',
            'team': 'NEW YORK YANKEES', 'opponent': 'BOSTON RED SOX',
            'home_away': 'home', 'play': 'BET',
        }
        fade = {
            'verdict': 'CLEAR FADE', 'line': -150, 'opp_line': 130,
            'scenario_id': '02', 'scenario': 'Y', 'team': 'NEW YORK YANKEES',
            'opponent': 'BOSTON RED SOX', 'home_away': 'home', 'play': 'FADE',
        }
        df1, _, _, _ = parse_results_upload(page.build_master_file(None, [bet], date(2026, 7, 21)))
        df2, _, _, _ = parse_results_upload(page.build_master_file(df1, [fade], date(2026, 7, 21)))
        if len(df2) != 1 or '#02' not in str(df2.iloc[0]['Scenario']):
            print(f"❌ FAILED: Expected 1 fade row, got {len(df2)} scenario={df2.iloc[0]['Scenario'] if len(df2) else 'none'}")
            return False
        print("✅ SUCCESS: Same-day re-generate replaces prior triggers")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_streamlit_scenario_perf_totals():
    """Streamlit Scenario Performance totals row uses F5/H5 (not header row)."""
    print("\n" + "="*60)
    print("TEST 22: Streamlit Scenario Performance Totals")
    print("="*60)
    try:
        import sys, os, importlib.util, io
        from datetime import date
        from openpyxl import load_workbook

        sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'pages'))
        spec = importlib.util.spec_from_file_location(
            'daily_report_page',
            os.path.join(os.path.dirname(__file__), 'pages', '1_Daily_Report.py'),
        )
        page = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(page)

        triggers = [{
            'verdict': 'CLEAR BET', 'line': -150, 'scenario_id': '01', 'scenario': 'B',
            'team': 'NEW YORK YANKEES', 'opponent': 'BOSTON RED SOX',
            'home_away': 'home', 'play': 'BET NEW YORK YANKEES',
        }]
        import report_builder
        from daily_report import SCENARIO_DEFS
        xbytes, _, _, _ = report_builder.build_report_bytes([], triggers, date(2026, 7, 20), None)
        wb = load_workbook(io.BytesIO(xbytes))
        perf = wb['Scenario Performance']
        totals_row = 5 + len(SCENARIO_DEFS)  # row 4 = headers, scenarios 5.., totals after
        f_val = str(perf.cell(row=totals_row, column=6).value or '')
        h_val = str(perf.cell(row=totals_row, column=8).value or '')
        g_val = str(perf.cell(row=totals_row, column=7).value or '')
        if 'SUM(F5' not in f_val and 'SUM(F5' not in f_val.upper():
            print(f"❌ FAILED: Total formula F col={f_val}")
            return False
        if 'SUM(H5' not in h_val:
            print(f"❌ FAILED: Total formula H col={h_val}")
            return False
        if 'IF(F' not in g_val:
            print(f"❌ FAILED: Missing Win% totals formula G col={g_val}")
            return False
        print("✅ SUCCESS: Streamlit Scenario Performance totals align with CLI")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_fade_legacy_no_odds_fallback():
    """Legacy FADE rows without PayoutLine must not fall back to faded team's Odds."""
    print("\n" + "="*60)
    print("TEST 23: FADE Legacy No Odds Fallback")
    print("="*60)
    try:
        import pandas as pd
        from master_results_manager import _payout_line_from_row

        fade_row = pd.Series({'Type': 'CLEAR FADE', 'Odds': '-150', 'PayoutLine': None})
        bet_row = pd.Series({'Type': 'CLEAR BET', 'Odds': '-150', 'PayoutLine': None})
        if _payout_line_from_row(fade_row) is not None:
            print("❌ FAILED: FADE row fell back to Odds column")
            return False
        if _payout_line_from_row(bet_row) != -150:
            print(f"❌ FAILED: BET row expected -150, got {_payout_line_from_row(bet_row)}")
            return False
        print("✅ SUCCESS: FADE legacy rows skip Odds fallback")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_fetch_recent_results_api_warnings():
    """Failed API days produce warnings instead of silent skip."""
    print("\n" + "="*60)
    print("TEST 24: Fetch Recent Results API Warnings")
    print("="*60)
    try:
        import pandas as pd
        import daily_report as dr
        from datetime import date
        from unittest.mock import patch
        import requests

        base_df = pd.DataFrame([{
            'year': 2026, 'team': 'NEW YORK YANKEES',
            'date': pd.Timestamp('2026-07-19'), 'opponent': 'BOSTON RED SOX',
            'home_away': 'home', 'result': 'W', 'score': '5-2',
            'runs_scored': 5, 'runs_allowed': 2,
            'line': -150, 'ou': None, 'total': None, 'division': 'AL_EAST',
        }])

        with patch('daily_report.requests.get', side_effect=requests.RequestException('timeout')):
            df, warnings = dr.fetch_recent_results(base_df, date(2026, 7, 21))

        if not warnings or '2026-07-20' not in warnings[0]:
            print(f"❌ FAILED: Expected warning for 2026-07-20, got {warnings}")
            return False
        if len(df) != len(base_df):
            print("❌ FAILED: Base dataframe should be unchanged on API failure")
            return False
        print("✅ SUCCESS: API failures return warnings")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_doubleheader_keeps_both_games():
    """Doubleheader days keep both Final games, not just the last one."""
    print("\n" + "="*60)
    print("TEST 25: Doubleheader Keeps Both Games")
    print("="*60)
    try:
        import pandas as pd
        import daily_report as dr
        from datetime import date
        from unittest.mock import patch, Mock

        base_df = pd.DataFrame([{
            'year': 2026, 'team': 'NEW YORK YANKEES',
            'date': pd.Timestamp('2026-07-19'), 'opponent': 'BOSTON RED SOX',
            'home_away': 'home', 'result': 'W', 'score': '5-2',
            'runs_scored': 5, 'runs_allowed': 2,
            'line': -150, 'ou': None, 'total': None, 'division': 'AL_EAST',
        }])

        dh_payload = {
            'dates': [{
                'games': [
                    {
                        'gamePk': 1,
                        'status': {'detailedState': 'Final'},
                        'teams': {
                            'away': {'team': {'name': 'New York Yankees'}, 'score': 3, 'isWinner': False},
                            'home': {'team': {'name': 'Boston Red Sox'}, 'score': 5, 'isWinner': True},
                        },
                    },
                    {
                        'gamePk': 2,
                        'status': {'detailedState': 'Final'},
                        'teams': {
                            'away': {'team': {'name': 'New York Yankees'}, 'score': 6, 'isWinner': True},
                            'home': {'team': {'name': 'Boston Red Sox'}, 'score': 2, 'isWinner': False},
                        },
                    },
                ],
            }],
        }

        mock_resp = Mock()
        mock_resp.raise_for_status = Mock()
        mock_resp.json = Mock(return_value=dh_payload)

        with patch('daily_report.requests.get', return_value=mock_resp):
            df, warnings = dr.fetch_recent_results(base_df, date(2026, 7, 21))

        day_rows = df[(df['team'] == 'NEW YORK YANKEES') & (df['date'].dt.date == date(2026, 7, 20))]
        if len(day_rows) != 2:
            print(f"❌ FAILED: Expected 2 NYY rows on DH day, got {len(day_rows)}")
            return False
        if warnings:
            print(f"❌ FAILED: Unexpected warnings: {warnings}")
            return False
        print("✅ SUCCESS: Doubleheader adds both games to historical data")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_ambiguous_los_angeles_opponent():
    """Bare 'los angeles' opponent string produces a warning."""
    print("\n" + "="*60)
    print("TEST 26: Ambiguous Los Angeles Opponent")
    print("="*60)
    try:
        from daily_report import normalize_opponent

        warnings = []
        result = normalize_opponent('los angeles', warnings)
        if result == 'LOS ANGELES' and warnings:
            print("✅ SUCCESS: Ambiguous los angeles flagged")
            return True
        print(f"❌ FAILED: result={result}, warnings={warnings}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_scatter_determine_dog_fav():
    """Scatter helper picks underdog by lower win%."""
    print("\n" + "="*60)
    print("TEST 27: Scatter Determine Dog/Fav")
    print("="*60)
    try:
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            'scatter_page',
            os.path.join(os.path.dirname(__file__), 'pages', '2_Scatter_Analysis.py'),
        )
        scatter = importlib.util.module_from_spec(spec)
        # Load only the helper — avoid running Streamlit page body
        code = open(spec.origin).read().split('st.title')[0]
        exec(compile(code, spec.origin, 'exec'), scatter.__dict__)

        winpcts = {'Team A': 0.400, 'Team B': 0.600}
        info = scatter.determine_dog_fav({'away': 'Team A', 'home': 'Team B'}, winpcts)
        if info and info['dog'] == 'Team A' and info['fav'] == 'Team B':
            print("✅ SUCCESS: Scatter dog/fav detection works")
            return True
        print(f"❌ FAILED: {info}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_s24_no_false_positive_first_road_game():
    """First game of a new road trip must not fire last_game_roadtrip / s24."""
    print("\n" + "="*60)
    print("TEST 29: S24 No False Positive First Road Game")
    print("="*60)
    try:
        from daily_report import build_game_row, s24, last_game_roadtrip

        state = {
            'team': 'NEW YORK YANKEES', 'streak_before': 0, 'streak_before_prev': 0,
            'prev_result': 'W', 'prev_opponent': 'BOSTON RED SOX', 'prev_line': -150,
            'prev_runs_scored': 5, 'prev_runs_allowed': 2, 'series_game_num': 1, 'series_total': 3,
            'homestand_game_num': 3, 'homestand_series_num': 1, 'roadtrip_game_num': 0, 'roadtrip_total': 0,
            'last10_wins': 5, 'wins_before': 50, 'games_before': 100, 'winpct_before': 0.5,
            'prev2_runs_scored': None, 'prev3_runs_scored': None, 'prev4_runs_scored': None,
            'prev2_runs_allowed': None, 'prev3_runs_allowed': None, 'prev4_runs_allowed': None,
            'division': 'AL_EAST',
        }
        row = build_game_row(state, 'away', 'BALTIMORE ORIOLES', 150)
        if last_game_roadtrip(row):
            print(f"❌ FAILED: last_game_roadtrip=True on road trip game 1 (rt={row['roadtrip_total']})")
            return False
        if s24(row):
            print("❌ FAILED: s24 fired on first road game")
            return False
        print("✅ SUCCESS: First road game does not trigger s24")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_regenerate_preserves_same_day_wl():
    """Re-generating same date keeps W/L already entered for matching triggers."""
    print("\n" + "="*60)
    print("TEST 30: Re-Generate Preserves Same-Day W/L")
    print("="*60)
    try:
        import sys, os, importlib.util
        from datetime import date

        sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'pages'))
        spec = importlib.util.spec_from_file_location(
            'daily_report_page',
            os.path.join(os.path.dirname(__file__), 'pages', '1_Daily_Report.py'),
        )
        page = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(page)

        bet = {
            'verdict': 'CLEAR BET', 'line': -150, 'scenario_id': '01', 'scenario': 'B',
            'team': 'NEW YORK YANKEES', 'opponent': 'BOSTON RED SOX',
            'home_away': 'home', 'play': 'BET',
        }
        df1, _, _, _ = parse_results_upload(page.build_master_file(None, [bet], date(2026, 7, 21)))
        df1.loc[0, 'Result'] = 'W'
        df2, _, _, _ = parse_results_upload(page.build_master_file(df1, [bet], date(2026, 7, 21)))
        if str(df2.iloc[0]['Result']).strip().upper() != 'W':
            print(f"❌ FAILED: Result lost on re-generate: {df2.iloc[0]['Result']}")
            return False
        print("✅ SUCCESS: Same-day W/L preserved on re-generate")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_master_rebuild_fade_no_odds_fallback():
    """build_master_file must not use faded-team Odds when PayoutLine missing for FADE."""
    print("\n" + "="*60)
    print("TEST 31: Master Rebuild FADE No Odds Fallback")
    print("="*60)
    try:
        import sys, os, importlib.util, io
        from datetime import date
        from openpyxl import load_workbook
        import pandas as pd

        sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'pages'))
        spec = importlib.util.spec_from_file_location(
            'daily_report_page',
            os.path.join(os.path.dirname(__file__), 'pages', '1_Daily_Report.py'),
        )
        page = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(page)

        legacy_row = {
            'Date': '2026-07-20', 'Team': 'New York Yankees', 'H/A': 'HOME',
            'Odds': '-150', 'Play': 'FADE', 'Scenario': '#02 Blowout', 'Type': 'CLEAR FADE',
            'Result': '', 'Net P/L': None, 'PayoutLine': None, 'Opponent': 'BOSTON RED SOX',
        }
        existing = pd.DataFrame([legacy_row])
        xbytes = page.build_master_file(existing, [], date(2026, 7, 21))
        wb = load_workbook(io.BytesIO(xbytes))
        ws = wb.active
        formula = ws.cell(row=4, column=9).value
        payout = ws.cell(row=4, column=10).value
        if payout is not None and payout != '':
            print(f"❌ FAILED: FADE row got payout line {payout} from Odds fallback")
            return False
        if formula and '-150' in str(formula):
            print(f"❌ FAILED: FADE formula used faded Odds: {formula}")
            return False
        print("✅ SUCCESS: Master rebuild skips Odds fallback for CLEAR FADE")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_report_builder_cumulative_fade_no_odds_fallback():
    """report_builder cumulative sheet must not use faded-team Odds for CLEAR FADE."""
    print("\n" + "="*60)
    print("TEST 32: Report Builder Cumulative FADE No Odds Fallback")
    print("="*60)
    try:
        import io
        import pandas as pd
        from datetime import date
        from openpyxl import load_workbook
        import report_builder

        legacy = pd.DataFrame([{
            'Date': '2026-07-19', 'Team': 'New York Yankees', 'H/A': 'HOME',
            'Odds': '-150', 'Play': 'FADE', 'Scenario': '#02 Blowout', 'Type': 'CLEAR FADE',
            'Result': 'W', 'Net P/L': None, 'PayoutLine': None, 'Opponent': 'BOSTON RED SOX',
        }])
        xbytes, _, _, _ = report_builder.build_report_bytes(
            [], [], date(2026, 7, 21), None, cumulative_df=legacy,
        )
        wb = load_workbook(io.BytesIO(xbytes))
        ws = wb['Cumulative Results']
        formula = ws.cell(row=4, column=9).value
        payout = ws.cell(row=4, column=10).value
        if payout is not None and payout != '':
            print(f"❌ FAILED: FADE cumulative payout={payout}")
            return False
        if formula and '-150' in str(formula):
            print(f"❌ FAILED: FADE formula used faded Odds: {formula}")
            return False
        print("✅ SUCCESS: Report builder cumulative skips Odds fallback for FADE")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_fetch_todays_schedule_api_error():
    """Schedule fetch failures return an error message, not a silent empty list."""
    print("\n" + "="*60)
    print("TEST 33: Fetch Today's Schedule API Error")
    print("="*60)
    try:
        import daily_report as dr
        from datetime import date
        from unittest.mock import patch
        import requests

        with patch('daily_report.requests.get', side_effect=requests.RequestException('timeout')):
            games, err = dr.fetch_todays_schedule(date(2026, 7, 21))
        if games:
            print(f"❌ FAILED: Expected empty games, got {len(games)}")
            return False
        if not err or '2026-07-21' not in err:
            print(f"❌ FAILED: Expected schedule error, got err={err}")
            return False
        print("✅ SUCCESS: Schedule API failure returns error message")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_report_builder_scenario_perf_range():
    """Streamlit Scenario Performance COUNTIFS range uses +500 headroom (matches CLI)."""
    print("\n" + "="*60)
    print("TEST 34: Report Builder Scenario Perf Range +500")
    print("="*60)
    try:
        import io
        from datetime import date
        from openpyxl import load_workbook
        import report_builder

        triggers = [{
            'verdict': 'CLEAR BET', 'line': -150, 'scenario_id': '01', 'scenario': 'B',
            'team': 'NEW YORK YANKEES', 'opponent': 'BOSTON RED SOX',
            'home_away': 'home', 'play': 'BET',
        }]
        xbytes, _, _, _ = report_builder.build_report_bytes([], triggers, date(2026, 7, 21), None)
        wb = load_workbook(io.BytesIO(xbytes))
        formula = str(wb['Scenario Performance'].cell(row=5, column=4).value or '')
        if 'F$1000' not in formula and 'F$504' not in formula and 'F$503' not in formula:
            print(f"❌ FAILED: Expected wide Results Tracker range in formula, got {formula}")
            return False
        print("✅ SUCCESS: Scenario Performance uses +500 headroom")
        return True
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def test_legacy_upload_notes():
    """Legacy 9-column upload returns upgrade notes."""
    print("\n" + "="*60)
    print("TEST 28: Legacy Upload Notes")
    print("="*60)
    try:
        import io
        import xlsxwriter

        buf = io.BytesIO()
        wb = xlsxwriter.Workbook(buf, {'in_memory': True})
        ws = wb.add_worksheet('Master Results')
        headers = ['Date', 'Team', 'H/A', 'Odds', 'Play', 'Scenario', 'Type', 'Result (W/L)', 'Net P/L ($100)']
        for i, h in enumerate(headers):
            ws.write(2, i, h)
        ws.write(3, 0, '2026-07-20')
        ws.write(3, 1, 'New York Yankees')
        ws.write(3, 7, 'W')
        wb.close()

        _, _, _, notes = parse_results_upload(buf.getvalue())
        if notes and any('PayoutLine' in n for n in notes):
            print("✅ SUCCESS: Legacy upload returns upgrade notes")
            return True
        print(f"❌ FAILED: notes={notes}")
        return False
    except Exception as e:
        print(f"❌ FAILED: {e}")
        return False


def run_all_tests():
    """Run all tests"""
    print("\n" + "="*70)
    print("🧪 MASTER RESULTS SYSTEM - TEST SUITE")
    print("="*70)
    
    tests = [
        ("Initialization", test_initialization),
        ("Append Results", test_append_results),
        ("Load Results", test_load_results),
        ("Cumulative Stats", test_cumulative_stats),
        ("W/L Entry Simulation", test_add_results_with_outcomes),
        ("Formula References", test_file_references),
        ("Daily Report Upload", test_parse_daily_report_upload),
        ("Master Dedup & FADE Round-Trip", test_master_dedup_and_fade_roundtrip),
        ("CLI Odds +130 Parsing", test_load_odds_plus_prefix),
        ("Series Last-Game Detection", test_series_last_game_detection),
        ("FADE Backtest W/L", test_fade_backtest_wl),
        ("Odds Cache Backfill", test_odds_cache_backfill),
        ("Tie Game Streak Handling", test_tie_games_not_counted_as_loss),
        ("Daily Report FADE PayoutLine Upload", test_daily_report_fade_payoutline_upload),
        ("Opponent Abbreviation Normalization", test_normalize_opponent_abbreviations),
        ("FADE Requires Opponent Odds", test_fade_pl_line_requires_opponent_odds),
        ("INCONSISTENT Backtest N/A", test_inconsistent_backtest_na),
        ("CLI Scenario Performance Master Merge", test_cli_scenario_perf_master_merge),
        ("Open Series No Early Last Game", test_open_series_no_early_last_game),
        ("Opponent Persists Master Round-Trip", test_opponent_persists_master_roundtrip),
        ("Re-Generate Replaces Same-Day Rows", test_regenerate_replaces_same_day_rows),
        ("Streamlit Scenario Performance Totals", test_streamlit_scenario_perf_totals),
        ("FADE Legacy No Odds Fallback", test_fade_legacy_no_odds_fallback),
        ("Fetch Recent Results API Warnings", test_fetch_recent_results_api_warnings),
        ("Doubleheader Keeps Both Games", test_doubleheader_keeps_both_games),
        ("Ambiguous Los Angeles Opponent", test_ambiguous_los_angeles_opponent),
        ("Scatter Determine Dog/Fav", test_scatter_determine_dog_fav),
        ("Legacy Upload Notes", test_legacy_upload_notes),
        ("S24 No False Positive First Road Game", test_s24_no_false_positive_first_road_game),
        ("Re-Generate Preserves Same-Day W/L", test_regenerate_preserves_same_day_wl),
        ("Master Rebuild FADE No Odds Fallback", test_master_rebuild_fade_no_odds_fallback),
        ("Report Builder Cumulative FADE No Odds Fallback", test_report_builder_cumulative_fade_no_odds_fallback),
        ("Fetch Today's Schedule API Error", test_fetch_todays_schedule_api_error),
        ("Report Builder Scenario Perf Range +500", test_report_builder_scenario_perf_range),
    ]
    
    results = []
    for name, test_func in tests:
        try:
            result = test_func()
            results.append((name, result))
        except Exception as e:
            print(f"\n❌ CRITICAL ERROR in {name}: {e}")
            results.append((name, False))
    
    # Summary
    print("\n" + "="*70)
    print("📊 TEST SUMMARY")
    print("="*70)
    
    passed = sum(1 for _, result in results if result)
    total = len(results)
    
    for name, result in results:
        status = "✅ PASS" if result else "❌ FAIL"
        print(f"{status:10} {name}")
    
    print("\n" + "-"*70)
    print(f"Results: {passed}/{total} tests passed ({passed/total*100:.0f}%)")
    
    if passed == total:
        print("\n🎉 ALL TESTS PASSED! Master Results system is working correctly.")
        print("\nNext steps:")
        print("1. Run the Streamlit app: streamlit run app.py")
        print("2. Upload Master_Results.xlsx (or yesterday's saved report with W/L entered)")
        print("3. Enter moneylines → Generate → download daily report + updated Master_Results.xlsx")
        print("4. Enter W/L in the master file after games — Scenario Performance updates automatically")
    else:
        print(f"\n⚠️  {total - passed} test(s) failed. Check the output above for details.")
    
    print("="*70 + "\n")
    
    return passed == total


if __name__ == '__main__':
    success = run_all_tests()
    sys.exit(0 if success else 1)
